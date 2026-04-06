#!/usr/bin/env python3
"""
claude_code_agent_ecosystem.py
Production multi-agent system for First Genesis with:
  - Stage gate framework (5 standard gates across agents)
  - Email notifications via Outlook SMTP (per-gate recipients)
  - Approval workflow (human replies, agent resumes automatically)
  - Cost Controller (track spend, enforce limits)
  - Budget Enforcer (hard stops before overspending)
  - Hallucination Guard (detect & prevent false claims)
  - Workflow persistence (SQLite state machine)

Configuration:
    export ANTHROPIC_API_KEY="sk-..."
    export OUTLOOK_SENDER="your-email@yourdomain.com"
    export OUTLOOK_PASSWORD="your-password"

Usage:
    python claude_code_agent_ecosystem.py run_pm_agent
    python claude_code_agent_ecosystem.py check_approvals
    python claude_code_agent_ecosystem.py resume_workflows
    python claude_code_agent_ecosystem.py process_approval --workflow-id <id> --approver <email> --decision <approved|rejected>
    python claude_code_agent_ecosystem.py budget_status
    python claude_code_agent_ecosystem.py audit_hallucinations
"""
import anthropic
import json
import os
import re
import sys
import uuid
import threading
import queue
import csv
import io
from datetime import datetime, timedelta
from typing import Optional, Dict, List, Tuple
from dataclasses import dataclass, asdict
from abc import ABC, abstractmethod
import sqlite3
import hashlib
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from enum import Enum

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)-8s %(message)s',
    handlers=[
        logging.FileHandler(os.environ.get("FG_LOG_PATH", "/home/claude/fg_agents.log")),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================================
# ENVIRONMENT CONFIGURATION
# ============================================================================
# Set FG_ENV=lab for lab/test mode (cheaper model, emails suppressed, relaxed budgets)
# Set FG_ENV=production (default) for full guardrails
FG_ENV = os.environ.get("FG_ENV", "production").lower()
IS_LAB_MODE = FG_ENV == "lab"

# Model selection: lab uses haiku (cheap), production uses opus
ACTIVE_MODEL = "claude-haiku-4-5-20251001" if IS_LAB_MODE else "claude-opus-4-6"

# Budget multiplier: lab mode has 10x relaxed caps to allow testing
LAB_BUDGET_MULTIPLIER = 10.0 if IS_LAB_MODE else 1.0

if IS_LAB_MODE:
    logger.warning("=" * 60)
    logger.warning("RUNNING IN LAB MODE: emails suppressed, haiku model, relaxed budgets")
    logger.warning("Set FG_ENV=production before go-live")
    logger.warning("=" * 60)

# ============================================================================
# CONFIGURATION: TEMPLATES
# ============================================================================
TEMPLATES = {
    "pm_agent": {
        "project_charter": """
PROJECT CHARTER
Title: {project_name}
Client: {client_name}
Budget: ${budget}
Timeline: {timeline} weeks
Scope: {scope}
Success Criteria:
  - {criteria_1}
  - {criteria_2}
Risks:
  - {risk_1}: Mitigation: {mitigation_1}
Created: {created_date}
""",
        "wbs": """
WORK BREAKDOWN STRUCTURE
Project: {project_name}
Phase 1: Project Initiation
  - Task 1.1: Kickoff meeting
  - Task 1.2: Baseline establishment
Phase 2: Design & Planning
  - Task 2.1: Design documentation
  - Task 2.2: Planning sessions
Phase 3: Execution
  - Task 3.1: Implementation
  - Task 3.2: Testing
Phase 4: Closure
  - Task 4.1: Handoff
  - Task 4.2: Lessons learned
""",
        "kickoff_checklist": """
PROJECT KICKOFF CHECKLIST
Project: {project_name}
[ ] Charter approved by stakeholders
[ ] Team roles and responsibilities assigned
[ ] Project schedule communicated
[ ] Risks documented and owned
[ ] Communication plan established
[ ] Kickoff meeting scheduled
[ ] Success criteria understood
[ ] Budget approved
[ ] Resources allocated
""",
    },
    "ba_agent": {
        "requirements": """
REQUIREMENTS SPECIFICATION
Project: {project_name}
FUNCTIONAL REQUIREMENTS:
FR1: {requirement_1}
  Acceptance Criteria: {criteria_1}
  Priority: {priority_1}
FR2: {requirement_2}
  Acceptance Criteria: {criteria_2}
  Priority: {priority_2}
NON-FUNCTIONAL REQUIREMENTS:
NFR1: Performance - {performance_req}
NFR2: Security - {security_req}
NFR3: Usability - {usability_req}
ASSUMPTIONS:
  - {assumption_1}
  - {assumption_2}
CONSTRAINTS:
  - {constraint_1}
  - {constraint_2}
""",
        "traceability_matrix": """
REQUIREMENTS TRACEABILITY MATRIX
Project: {project_name}
ID  | Requirement | Design Doc | Test Case | Status
----|-------------|-----------|-----------|--------
FR1 | {req_1}     | DESIGN-1  | TEST-1   | {status_1}
FR2 | {req_2}     | DESIGN-2  | TEST-2   | {status_2}
NFR1| {req_3}     | DESIGN-3  | TEST-3   | {status_3}
""",
        "design_session_template": """
DESIGN SESSION NOTES
Project: {project_name}
Date: {session_date}
Attendees: {attendees}
DISCUSSION POINTS:
1. {topic_1}
   Decision: {decision_1}
   Owner: {owner_1}
2. {topic_2}
   Decision: {decision_2}
   Owner: {owner_2}
NEXT STEPS:
[ ] {action_1} - Owner: {owner_1} - Due: {due_date_1}
[ ] {action_2} - Owner: {owner_2} - Due: {due_date_2}
OPEN ITEMS:
- {open_item_1}
- {open_item_2}
""",
        "data_schema_extraction": """
DATA SCHEMA EXTRACTION REPORT
Project: {project_name}
Source Document: {filename}
Extracted By: BA Agent
Date: {extraction_date}

SOURCE SUMMARY:
  File Type:    {file_type}
  Row Count:    {row_count}
  PII Tokens Redacted: {pii_count}

INFERRED SCHEMA:
Field Name             | Inferred Type  | Nullable | Notes
-----------------------|----------------|----------|------------------------
{schema_rows}

DATA QUALITY:
  Completeness:   {completeness_pct}%
  Issues Found:   {issue_count}
  Issues:
{issues}

REQUIREMENTS EXTRACTED FROM DATA:
{data_requirements}

SME REVIEW REQUIRED FOR:
{sme_fields}

HUMAN_ONLY — Data Architect sign-off:
[HUMAN_ONLY]

Reviewed By: [HUMAN_ONLY]
""",
        "data_dependency_map": """
DATA DEPENDENCY MAP
Project: {project_name}
Date: {map_date}

UPSTREAM DEPENDENCIES (data this project consumes):
{upstream_deps}

DOWNSTREAM DEPENDENCIES (data this project produces):
{downstream_deps}

MISSING / UNRESOLVED DEPENDENCIES:
{missing_deps}

INTEGRATION POINTS: [HUMAN_ONLY]
TECHNICAL DATA FLOW: [HUMAN_ONLY]
""",
    },
    "qa_agent": {
        "qa_checklist": """
PRE-DELIVERY QA CHECKLIST
Project: {project_name}
Date: {review_date}
COMPLETENESS:
[ ] All requirements satisfied
[ ] All acceptance criteria met
[ ] Documentation complete
[ ] Test coverage adequate
[ ] Known issues documented
QUALITY:
[ ] No major defects found
[ ] Code quality acceptable
[ ] Performance acceptable
[ ] Security review passed
[ ] Accessibility standards met
SCOPE CREEP:
[ ] No scope changes outside approved RFP
[ ] Budget variance < 5%
[ ] Timeline variance < 10%
[ ] Team additions pre-approved
READINESS:
[ ] Customer expectations aligned
[ ] Support team prepared
[ ] Deployment plan ready
[ ] Rollback plan ready
OVERALL ASSESSMENT: {assessment}
Recommendation: READY FOR DELIVERY / NEEDS WORK
Reviewed By: {reviewer}
Approved By: {approver}
""",
        "scope_creep_detection": """
SCOPE CREEP DETECTION RULES
Project: {project_name}
Rule 1: Unplanned Features
  Trigger: New feature request not in original RFP
  Action: Alert PM, get approval before proceeding
Rule 2: Timeline Variance
  Trigger: Schedule change > 10%
  Action: Alert PM and customer
Rule 3: Budget Variance
  Trigger: Cost variance > 5%
  Action: Escalate to CEO if critical
Rule 4: Team Changes
  Trigger: Resource additions not pre-approved
  Action: Alert PM, verify budget impact
Rule 5: Requirement Changes
  Trigger: Requirements modified post-approval
  Action: Document change, get sign-off
CURRENT PROJECT STATUS:
Scope variance: {scope_variance}%
Timeline variance: {timeline_variance}%
Budget variance: {budget_variance}%
Status: {creep_status}
""",
    },
    "vendor_agent": {
        "sla_template": """
VENDOR/PARTNER SLA
Vendor: {vendor_name}
Engagement: {engagement_name}
SCOPE:
{scope}
DELIVERABLES:
  1. {deliverable_1} - Due: {due_date_1}
  2. {deliverable_2} - Due: {due_date_2}
  3. {deliverable_3} - Due: {due_date_3}
TIMELINE: {start_date} to {end_date} ({duration} weeks)
BUDGET: ${budget}
PERFORMANCE METRICS:
  - On-time delivery: 95%+ required
  - Quality score: 90%+ required
  - Communication: <24h response time
  - Escalation handling: <4 business hours
PAYMENT TERMS:
  {payment_terms}
TERMINATION CLAUSE:
  {termination_clause}
""",
        "scorecard_template": """
VENDOR PERFORMANCE SCORECARD
Vendor: {vendor_name}
Period: {period}
Engagement: {engagement_name}
PERFORMANCE METRICS:
Metric                 | Target | Actual | Status   | Notes
-----------------------|--------|--------|----------|----------
On-time delivery       | 95%    | {pct1} | {st1}    | {notes1}
Quality score          | 90%    | {pct2} | {st2}    | {notes2}
Avg response time      | <24h   | {time} | {st3}    | {notes3}
Budget variance        | <5%    | {var}  | {st4}    | {notes4}
Issue resolution time  | <4h    | {hours}| {st5}    | {notes5}
OVERALL RATING: {rating}/5.0
Status: ON_TRACK / NEEDS_IMPROVEMENT / AT_RISK
RECOMMENDATIONS:
  1. {recommendation_1}
  2. {recommendation_2}
Reviewed By: {reviewer}
Date: {review_date}
""",
    },
    "manager_agent": {
        "portfolio_dashboard": """
PORTFOLIO MANAGEMENT DASHBOARD
Generated: {timestamp}
PROJECT STATUS OVERVIEW:
Project              | Status        | Owner         | Risk Level | Days Active
---------------------|---------------|---------------|------------|----------
{project_1}          | {status_1}    | {owner_1}     | {risk_1}   | {days_1}
{project_2}          | {status_2}    | {owner_2}     | {risk_2}   | {days_2}
{project_3}          | {status_3}    | {owner_3}     | {risk_3}   | {days_3}
BUDGET SUMMARY:
Project              | Budget    | Spent     | Remaining | % Used
---------------------|-----------|-----------|-----------|--------
{project_1}          | ${b1}     | ${s1}     | ${r1}     | {pct1}%
{project_2}          | ${b2}     | ${s2}     | ${r2}     | {pct2}%
{project_3}          | ${b3}     | ${s3}     | ${r3}     | {pct3}%
TOTAL                | ${total_b}| ${total_s}| ${total_r}| {total_pct}%
PENDING APPROVALS ({approval_count}):
  - {workflow_1} (waiting {hours_1}h)
  - {workflow_2} (waiting {hours_2}h)
RISKS & BLOCKERS ({blocker_count}):
  CRITICAL: {critical_blocker}
  HIGH: {high_blocker}
NEXT ACTIONS:
  [ ] {action_1} - Owner: {owner} - Due: {due_date}
  [ ] {action_2} - Owner: {owner} - Due: {due_date}
METRICS:
  Total Projects: {total_projects}
  On Track: {on_track}
  At Risk: {at_risk}
  Avg Project Health: {health_score}%
Report Generated: {timestamp}
""",
    }
}

# ============================================================================
# PHASE 1 REVIEW TEMPLATES
# Agents do NOT create deliverables in Phase 1.
# They review, assess, and approve vendor-submitted documents.
# ============================================================================
REVIEW_TEMPLATES = {
    "pm_agent": {
        "scope_compliance_review": """\
SCOPE COMPLIANCE REVIEW — PM Agent
Deliverable ID: {deliverable_id}
Vendor: {vendor_name}
Type: {deliverable_type}
Reviewed: {review_date}

PROJECT BASELINE (frozen facts):
  Client:    Malcolm Goodwin
  Project:   AURA MVP
  Scope:     Silhouette technology + 3D mesh design + actor model
  Timeline:  3 months (deadline April 30 2026)
  Vendor:    Yubi

VENDOR DELIVERABLE TEXT:
{deliverable_text}

Review this deliverable and respond in JSON only — no preamble, no explanation.
Mark any field you cannot assess as "[NEEDS_SME_INPUT]".

{{
  "scope_compliance_score": <1-5>,
  "in_scope_items": ["..."],
  "out_of_scope_items": ["..."],
  "scope_gaps": ["..."],
  "timeline_risk": "LOW|MEDIUM|HIGH",
  "timeline_notes": "...",
  "budget_impact": "NONE|LOW|MEDIUM|HIGH",
  "budget_notes": "...",
  "recommendation": "APPROVE|FLAG|REJECT",
  "pm_notes": "...",
  "confidence": <1-5>
}}
""",
    },
    "ba_agent": {
        "requirements_compliance_review": """\
REQUIREMENTS COMPLIANCE REVIEW — BA Agent
Deliverable ID: {deliverable_id}
Vendor: {vendor_name}
Type: {deliverable_type}
Reviewed: {review_date}

PROJECT REQUIREMENTS CONTEXT:
  Project:   AURA MVP
  Client:    Malcolm Goodwin
  Core Scope: Silhouette technology + 3D mesh design + actor model
  Deadline:  April 30 2026

VENDOR DELIVERABLE TEXT:
{deliverable_text}

Assess this deliverable against AURA project requirements.
Fill facts only. Respond in JSON only — no preamble.

{{
  "requirements_coverage_score": <1-5>,
  "satisfied_requirements": ["..."],
  "unsatisfied_requirements": ["..."],
  "missing_acceptance_criteria": ["..."],
  "data_quality_notes": "...",
  "integration_concerns": ["..."],
  "recommendation": "APPROVE|FLAG|REJECT",
  "ba_notes": "...",
  "confidence": <1-5>
}}
""",
    },
    "qa_agent": {
        "quality_assessment_review": """\
QUALITY ASSESSMENT REVIEW — QA Agent
Deliverable ID: {deliverable_id}
Vendor: {vendor_name}
Type: {deliverable_type}
Reviewed: {review_date}

VENDOR DELIVERABLE TEXT:
{deliverable_text}

Assess quality, completeness, accuracy, and standards compliance.
Respond in JSON only — no preamble.

{{
  "quality_score": <1-5>,
  "completeness_pct": <0-100>,
  "defects_found": ["..."],
  "quality_standards_met": ["..."],
  "quality_standards_failed": ["..."],
  "risk_level": "LOW|MEDIUM|HIGH",
  "scope_creep_detected": true|false,
  "scope_creep_notes": "...",
  "recommendation": "APPROVE|FLAG|REJECT",
  "qa_notes": "...",
  "confidence": <1-5>
}}
""",
    },
}

# ============================================================================
# DATA MODELS
# ============================================================================
class StageGateName(Enum):
    CHARTER_APPROVAL = "charter_approval"
    REQUIREMENTS_APPROVAL = "requirements_approval"
    QA_AUDIT_APPROVAL = "qa_audit_approval"
    DELIVERY_APPROVAL = "delivery_approval"
    BUDGET_ESCALATION = "budget_escalation"
    DOCUMENT_CLEARANCE = "document_clearance"   # inbound PII review before LLM ingestion
    LAB_SIGN_OFF = "lab_sign_off"               # required before promoting lab workflow to production
    SME_REVIEW = "sme_review"                   # triggered when agent confidence < 3
    COMBINED_REVIEW = "combined_review"         # Phase 1: all agents reviewed vendor deliverable → SME

class WorkflowStatus(Enum):
    PENDING = "pending"
    APPROVAL_SENT = "approval_sent"
    APPROVED = "approved"
    REJECTED = "rejected"
    RESUMED = "resumed"
    COMPLETED = "completed"

@dataclass
class StageGate:
    name: StageGateName
    description: str
    approver_email: str
    cc_emails: List[str] = None
    auto_approve_if: Optional[str] = None
    require_comment: bool = False
    timeout_hours: int = 24

@dataclass
class WorkflowState:
    workflow_id: str
    agent_name: str
    project_name: str
    current_stage_gate: StageGateName
    status: WorkflowStatus
    content_pending_approval: str
    content_hash: str
    approval_email_sent_at: Optional[str] = None
    human_approver: Optional[str] = None
    human_feedback: Optional[str] = None
    approval_timestamp: Optional[str] = None
    next_step_after_approval: Optional[str] = None
    created_at: str = None
    updated_at: str = None

    def __post_init__(self):
        if self.created_at is None:
            self.created_at = datetime.now().isoformat()
        self.updated_at = datetime.now().isoformat()

@dataclass
class TokenCost:
    """Track token usage and cost (claude-opus-4-6 pricing)."""
    input_tokens: int
    output_tokens: int
    cache_creation_tokens: int = 0
    cache_read_tokens: int = 0
    model: str = "claude-opus-4-6"

    INPUT_COST_PER_1M = 5.00
    OUTPUT_COST_PER_1M = 25.00
    CACHE_WRITE_COST_PER_1M = 0.50
    CACHE_READ_COST_PER_1M = 0.50

    def total_cost_usd(self) -> float:
        input_cost = (self.input_tokens / 1_000_000) * self.INPUT_COST_PER_1M
        output_cost = (self.output_tokens / 1_000_000) * self.OUTPUT_COST_PER_1M
        cache_write = (self.cache_creation_tokens / 1_000_000) * self.CACHE_WRITE_COST_PER_1M
        cache_read = (self.cache_read_tokens / 1_000_000) * self.CACHE_READ_COST_PER_1M
        return input_cost + output_cost + cache_write + cache_read

    def __str__(self) -> str:
        return (f"In={self.input_tokens} Out={self.output_tokens} "
                f"CW={self.cache_creation_tokens} CR={self.cache_read_tokens} "
                f"Cost=${self.total_cost_usd():.4f}")

@dataclass
class AgentConfig:
    name: str
    budget_per_call_usd: float
    max_daily_calls: int
    max_daily_spend_usd: float
    priority: int
    description: str

@dataclass
class AgentCall:
    agent_name: str
    timestamp: str
    input_tokens: int
    output_tokens: int
    cache_created: int
    cache_read: int
    cost_usd: float
    status: str
    reason: str
    output_hash: str

# ============================================================================
# DASHBOARD: ENUMS & DATACLASSES
# ============================================================================
class AgentStatus(Enum):
    IDLE = "idle"
    THINKING = "thinking"
    WAITING_INPUT = "waiting_input"
    EXECUTING = "executing"
    WAITING_APPROVAL = "waiting_approval"
    COMPLETE = "complete"
    ERROR = "error"

class SkillLevel(Enum):
    NOVICE = 1
    INTERMEDIATE = 2
    ADVANCED = 3
    EXPERT = 4

class MessageType(Enum):
    INITIATE = "initiate"
    REQUEST_INPUT = "request_input"
    PROVIDE_OUTPUT = "provide_output"
    DELEGATE = "delegate"
    FEEDBACK = "feedback"
    ESCALATE = "escalate"

@dataclass
class AgentRecord:
    """Agent state and metadata for dashboard monitoring."""
    id: str
    name: str
    role: str
    status: AgentStatus
    current_task: Optional[str]
    skill_level: SkillLevel
    success_count: int
    error_count: int
    last_activity: str
    active_workflows: List[str]

    def __post_init__(self):
        if not self.id:
            self.id = f"{self.name}_{uuid.uuid4().hex[:8]}"

@dataclass
class AgentMessage:
    """Message between agents via the communication bus."""
    id: str
    from_agent: str
    to_agent: str
    message_type: MessageType
    content: Dict
    timestamp: str
    status: str  # pending, sent, acknowledged, completed

@dataclass
class AgentSkill:
    """Learned capability tracked per agent."""
    skill_id: str
    agent_name: str
    skill_name: str
    description: str
    success_count: int
    error_count: int
    avg_execution_time: float
    skill_level: SkillLevel
    last_used: str
    template: Dict

@dataclass
class WorkflowExecution:
    """Record of a complete workflow execution for pattern capture."""
    workflow_id: str
    agent_sequence: List[str]
    start_time: str
    end_time: str
    duration_seconds: float
    success: bool
    input_data: Dict
    output_data: Dict
    errors: List[str]
    approvals_needed: int
    approvals_completed: int
    cost_usd: float

# ============================================================================
# DATABASE
# ============================================================================
class WorkflowDatabase:
    """Unified SQLite database for workflows, approvals, agent calls, and hallucination flags."""

    def __init__(self, db_path: str = None):
        if db_path is None:
            db_path = os.environ.get("FG_DB_PATH", "/home/claude/fg_workflows.db")
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path, check_same_thread=False)
        self.cursor = self.conn.cursor()
        self._init_tables()

    def _init_tables(self):
        self.cursor.executescript("""
            CREATE TABLE IF NOT EXISTS workflows (
                workflow_id TEXT PRIMARY KEY,
                agent_name TEXT NOT NULL,
                project_name TEXT NOT NULL,
                current_stage_gate TEXT NOT NULL,
                status TEXT NOT NULL,
                content_pending_approval TEXT,
                content_hash TEXT,
                approval_email_sent_at TEXT,
                human_approver TEXT,
                human_feedback TEXT,
                approval_timestamp TEXT,
                next_step_after_approval TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS approvals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                workflow_id TEXT NOT NULL,
                stage_gate TEXT NOT NULL,
                approver_email TEXT NOT NULL,
                decision TEXT NOT NULL,
                feedback TEXT,
                decided_at TEXT NOT NULL,
                FOREIGN KEY(workflow_id) REFERENCES workflows(workflow_id)
            );

            CREATE TABLE IF NOT EXISTS agent_calls (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                workflow_id TEXT,
                agent_name TEXT NOT NULL,
                timestamp TEXT NOT NULL,
                input_tokens INTEGER NOT NULL,
                output_tokens INTEGER NOT NULL,
                cache_created INTEGER DEFAULT 0,
                cache_read INTEGER DEFAULT 0,
                cost_usd REAL NOT NULL,
                status TEXT NOT NULL,
                reason TEXT,
                output_hash TEXT,
                FOREIGN KEY(workflow_id) REFERENCES workflows(workflow_id)
            );

            CREATE TABLE IF NOT EXISTS hallucination_flags (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                agent_name TEXT NOT NULL,
                timestamp TEXT NOT NULL,
                flag_reason TEXT NOT NULL,
                output_snippet TEXT
            );

            CREATE TABLE IF NOT EXISTS frozen_facts (
                fact_key TEXT PRIMARY KEY,
                fact_value TEXT NOT NULL,
                added_by TEXT NOT NULL,
                created_at TEXT NOT NULL,
                active INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS bdr_documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                doc_id TEXT UNIQUE NOT NULL,
                filename TEXT NOT NULL,
                file_type TEXT NOT NULL,
                raw_row_count INTEGER DEFAULT 0,
                extracted_fields TEXT,
                redaction_map TEXT,
                clearance_workflow_id TEXT,
                status TEXT DEFAULT 'pending_clearance',
                ingested_at TEXT NOT NULL,
                cleared_at TEXT
            );

            CREATE TABLE IF NOT EXISTS vendor_deliverables (
                deliverable_id TEXT PRIMARY KEY,
                vendor_name TEXT NOT NULL,
                deliverable_type TEXT NOT NULL,
                filename TEXT,
                content_text TEXT,
                submitted_at TEXT NOT NULL,
                review_status TEXT DEFAULT 'pending',
                pm_assessment TEXT,
                pm_score INTEGER,
                ba_assessment TEXT,
                ba_score INTEGER,
                qa_assessment TEXT,
                qa_score INTEGER,
                combined_review TEXT,
                overall_score REAL,
                overall_risk TEXT,
                recommendation TEXT,
                sme_decision TEXT,
                sme_feedback TEXT,
                workflow_id TEXT,
                updated_at TEXT
            );
        """)
        self.conn.commit()
        logger.info("Database initialized")

    # ── Workflow state ────────────────────────────────────────────────────────

    def save_workflow_state(self, state: WorkflowState):
        self.cursor.execute("""
            INSERT OR REPLACE INTO workflows
            (workflow_id, agent_name, project_name, current_stage_gate, status,
             content_pending_approval, content_hash, approval_email_sent_at,
             human_approver, human_feedback, approval_timestamp,
             next_step_after_approval, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            state.workflow_id, state.agent_name, state.project_name,
            state.current_stage_gate.value, state.status.value,
            state.content_pending_approval, state.content_hash,
            state.approval_email_sent_at, state.human_approver,
            state.human_feedback, state.approval_timestamp,
            state.next_step_after_approval, state.created_at, state.updated_at
        ))
        self.conn.commit()
        logger.info(f"Workflow {state.workflow_id} saved: {state.status.value}")

    def get_workflow_state(self, workflow_id: str) -> Optional[WorkflowState]:
        self.cursor.execute("SELECT * FROM workflows WHERE workflow_id = ?", (workflow_id,))
        row = self.cursor.fetchone()
        if not row:
            return None
        return WorkflowState(
            workflow_id=row[0], agent_name=row[1], project_name=row[2],
            current_stage_gate=StageGateName(row[3]), status=WorkflowStatus(row[4]),
            content_pending_approval=row[5], content_hash=row[6],
            approval_email_sent_at=row[7], human_approver=row[8],
            human_feedback=row[9], approval_timestamp=row[10],
            next_step_after_approval=row[11], created_at=row[12], updated_at=row[13]
        )

    def get_pending_approvals(self) -> List[WorkflowState]:
        self.cursor.execute(
            "SELECT * FROM workflows WHERE status = ? ORDER BY updated_at ASC",
            (WorkflowStatus.APPROVAL_SENT.value,)
        )
        return [self._row_to_workflow(r) for r in self.cursor.fetchall()]

    def get_approved_workflows_ready_to_resume(self) -> List[WorkflowState]:
        self.cursor.execute("""
            SELECT * FROM workflows WHERE status = ? AND approval_timestamp IS NOT NULL
            ORDER BY approval_timestamp ASC
        """, (WorkflowStatus.APPROVED.value,))
        return [self._row_to_workflow(r) for r in self.cursor.fetchall()]

    def _row_to_workflow(self, r) -> WorkflowState:
        return WorkflowState(
            workflow_id=r[0], agent_name=r[1], project_name=r[2],
            current_stage_gate=StageGateName(r[3]), status=WorkflowStatus(r[4]),
            content_pending_approval=r[5], content_hash=r[6],
            approval_email_sent_at=r[7], human_approver=r[8],
            human_feedback=r[9], approval_timestamp=r[10],
            next_step_after_approval=r[11], created_at=r[12], updated_at=r[13]
        )

    # ── Approvals ─────────────────────────────────────────────────────────────

    def record_approval(self, workflow_id: str, stage_gate: StageGateName,
                        approver_email: str, decision: str, feedback: str = None):
        self.cursor.execute("""
            INSERT INTO approvals
            (workflow_id, stage_gate, approver_email, decision, feedback, decided_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (workflow_id, stage_gate.value, approver_email, decision, feedback,
              datetime.now().isoformat()))
        self.conn.commit()
        logger.info(f"Approval recorded for {workflow_id}: {decision}")

    # ── Agent calls ───────────────────────────────────────────────────────────

    def log_agent_call(self, call: AgentCall, workflow_id: str = None):
        self.cursor.execute("""
            INSERT INTO agent_calls
            (workflow_id, agent_name, timestamp, input_tokens, output_tokens,
             cache_created, cache_read, cost_usd, status, reason, output_hash)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (workflow_id, call.agent_name, call.timestamp, call.input_tokens,
              call.output_tokens, call.cache_created, call.cache_read,
              call.cost_usd, call.status, call.reason, call.output_hash))
        self.conn.commit()

    def get_today_spend(self) -> Tuple[float, int]:
        today = datetime.now().strftime("%Y-%m-%d")
        self.cursor.execute("""
            SELECT COALESCE(SUM(cost_usd), 0), COUNT(*) FROM agent_calls
            WHERE DATE(timestamp) = ?
              AND status NOT IN ('rejected_budget', 'rejected_hallucination')
        """, (today,))
        spend, calls = self.cursor.fetchone()
        return spend or 0.0, calls or 0

    def get_agent_spend_today(self, agent_name: str) -> Tuple[float, int]:
        today = datetime.now().strftime("%Y-%m-%d")
        self.cursor.execute("""
            SELECT COALESCE(SUM(cost_usd), 0), COUNT(*) FROM agent_calls
            WHERE DATE(timestamp) = ? AND agent_name = ? AND status = 'success'
        """, (today, agent_name))
        spend, calls = self.cursor.fetchone()
        return spend or 0.0, calls or 0

    def get_last_n_days_spend(self, days: int = 7) -> float:
        date_ago = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
        self.cursor.execute("""
            SELECT COALESCE(SUM(cost_usd), 0) FROM agent_calls
            WHERE DATE(timestamp) >= ? AND status = 'success'
        """, (date_ago,))
        return self.cursor.fetchone()[0] or 0.0

    # ── Hallucination flags ───────────────────────────────────────────────────

    def log_hallucination_flag(self, agent_name: str, reason: str, snippet: str):
        self.cursor.execute("""
            INSERT INTO hallucination_flags (agent_name, timestamp, flag_reason, output_snippet)
            VALUES (?, ?, ?, ?)
        """, (agent_name, datetime.now().isoformat(), reason, snippet[:500]))
        self.conn.commit()
        logger.warning(f"Hallucination flagged: {agent_name} - {reason}")

# ============================================================================
# EMAIL GATEWAY (Outlook SMTP)
# ============================================================================
# EMAIL GATEWAY — Microsoft Graph API (primary) + SMTP (fallback)
#
# Provider selection (set in .env):
#   EMAIL_PROVIDER=graph   → Microsoft Graph API via OAuth2 client credentials
#   EMAIL_PROVIDER=smtp    → Outlook SMTP (legacy, requires app password)
#   EMAIL_PROVIDER=none    → Disable email (gates log only — useful in lab mode)
#
# Graph API env vars:
#   MS_TENANT_ID       Azure AD tenant ID (Directory ID)
#   MS_CLIENT_ID       App registration client ID
#   MS_CLIENT_SECRET   App registration client secret
#   MS_SENDER_EMAIL    Email address the agent sends FROM (kphipps@firstgenesis.com)
#
# SMTP env vars (fallback):
#   OUTLOOK_SENDER     Sender email
#   OUTLOOK_PASSWORD   Outlook app-specific password
# ============================================================================

import urllib.request
import urllib.parse
import urllib.error

class GraphEmailGateway:
    """Send approval emails via Microsoft Graph API (OAuth2 client credentials).

    Requires an Azure AD app registration with Mail.Send application permission
    and admin consent granted. No user login or app passwords needed.

    One-time Azure setup (5 minutes):
      1. portal.azure.com → Azure Active Directory → App registrations → New
      2. Name: FG-Agent-System  |  Account type: Single tenant
      3. Note: Application (client) ID  and  Directory (tenant) ID
      4. Certificates & secrets → New client secret → copy value
      5. API permissions → Add → Microsoft Graph → Application → Mail.Send
      6. Click "Grant admin consent for [tenant]"
    """

    GRAPH_TOKEN_URL = "https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    GRAPH_SEND_URL  = "https://graph.microsoft.com/v1.0/users/{sender}/sendMail"
    GRAPH_SCOPE     = "https://graph.microsoft.com/.default"

    def __init__(self):
        self.tenant_id     = os.environ.get("MS_TENANT_ID")
        self.client_id     = os.environ.get("MS_CLIENT_ID")
        self.client_secret = os.environ.get("MS_CLIENT_SECRET")
        self.sender_email  = os.environ.get("MS_SENDER_EMAIL", "kphipps@firstgenesis.com")
        self._token: Optional[str] = None
        self._token_expiry: float = 0.0

        missing = [k for k, v in {
            "MS_TENANT_ID": self.tenant_id,
            "MS_CLIENT_ID": self.client_id,
            "MS_CLIENT_SECRET": self.client_secret,
        }.items() if not v]

        if missing:
            logger.warning(f"Graph API email not configured. Missing: {missing}")
            self.enabled = False
        else:
            self.enabled = True
            logger.info(f"Graph API email gateway ready (sender: {self.sender_email})")

    def _get_token(self) -> Optional[str]:
        """Fetch or reuse an OAuth2 access token (cached, auto-refreshes)."""
        import time
        if self._token and time.time() < self._token_expiry - 60:
            return self._token

        url = self.GRAPH_TOKEN_URL.format(tenant_id=self.tenant_id)
        payload = urllib.parse.urlencode({
            "grant_type":    "client_credentials",
            "client_id":     self.client_id,
            "client_secret": self.client_secret,
            "scope":         self.GRAPH_SCOPE,
        }).encode()

        try:
            req = urllib.request.Request(url, data=payload, method="POST")
            req.add_header("Content-Type", "application/x-www-form-urlencoded")
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode())
            import time as _time
            self._token = data["access_token"]
            self._token_expiry = _time.time() + data.get("expires_in", 3600)
            logger.info("Graph API token acquired")
            return self._token
        except Exception as e:
            logger.error(f"Graph API token fetch failed: {e}")
            return None

    def _build_email_body(self, workflow_id: str, stage_gate: "StageGate",
                          agent_name: str, project_name: str,
                          content_summary: str, content_detail: str) -> str:
        return f"""APPROVAL REQUEST — First Genesis Agent System
{'='*60}
Agent:        {agent_name}
Project:      {project_name}
Stage Gate:   {stage_gate.description}
Workflow ID:  {workflow_id}
Sent:         {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} UTC

SUMMARY:
{content_summary}

CONTENT PREVIEW:
{content_detail[:600]}...

{'='*60}
ACTION REQUIRED — reply with one of:
  APPROVED   → agent proceeds to next step
  REJECTED   → workflow paused, no further action taken

You may include feedback after your decision word.
Auto-escalates in {stage_gate.timeout_hours} hour(s) if no reply received.
{'='*60}
This is an automated message from the FG Sales Agent System.
"""

    def send_approval_request(self, workflow_id: str, stage_gate: "StageGate",
                              agent_name: str, project_name: str,
                              content_summary: str, content_detail: str) -> bool:
        if not self.enabled:
            logger.warning(f"Graph email disabled. Approval for {workflow_id} not sent.")
            return False

        token = self._get_token()
        if not token:
            logger.error(f"Cannot send email for {workflow_id}: no access token")
            return False

        subject = (f"[Action Required] {stage_gate.description} — "
                   f"{project_name} ({agent_name})")
        body = self._build_email_body(
            workflow_id, stage_gate, agent_name, project_name,
            content_summary, content_detail
        )

        # Build recipient list
        to_recipients = [{"emailAddress": {"address": stage_gate.approver_email}}]
        cc_recipients = [
            {"emailAddress": {"address": addr}}
            for addr in (stage_gate.cc_emails or [])
        ]

        payload = json.dumps({
            "message": {
                "subject": subject,
                "body": {"contentType": "Text", "content": body},
                "toRecipients": to_recipients,
                "ccRecipients": cc_recipients,
            },
            "saveToSentItems": True
        }).encode()

        url = self.GRAPH_SEND_URL.format(sender=urllib.parse.quote(self.sender_email))
        try:
            req = urllib.request.Request(url, data=payload, method="POST")
            req.add_header("Authorization", f"Bearer {token}")
            req.add_header("Content-Type", "application/json")
            with urllib.request.urlopen(req, timeout=15) as resp:
                # Graph returns 202 Accepted on success (no body)
                success = resp.status == 202
            if success:
                logger.info(
                    f"Graph email sent for {workflow_id} → {stage_gate.approver_email}"
                )
            return success
        except urllib.error.HTTPError as e:
            err_body = e.read().decode()[:300]
            logger.error(f"Graph email HTTP {e.code} for {workflow_id}: {err_body}")
            return False
        except Exception as e:
            logger.error(f"Graph email failed for {workflow_id}: {e}")
            return False

    def parse_approval_response(self, email_subject: str, email_body: str) -> Tuple[str, str]:
        body_upper = email_body.upper()
        if "APPROVED" in body_upper:
            decision = "approved"
        elif "REJECTED" in body_upper:
            decision = "rejected"
        else:
            decision = "clarification_needed"
        return decision, email_body.strip()


class SmtpEmailGateway:
    """Send approval emails via Outlook SMTP (legacy fallback).
    Requires an Outlook app-specific password. Use GraphEmailGateway instead
    for Microsoft 365 business accounts.
    """

    def __init__(self):
        self.sender_email  = os.environ.get("OUTLOOK_SENDER")
        self.sender_password = os.environ.get("OUTLOOK_PASSWORD")
        self.smtp_server = "smtp.office365.com"
        self.smtp_port   = 587

        if not self.sender_email or not self.sender_password:
            logger.warning("SMTP email not configured (OUTLOOK_SENDER / OUTLOOK_PASSWORD missing).")
            self.enabled = False
        else:
            self.enabled = True
            logger.info(f"SMTP email gateway ready (sender: {self.sender_email})")

    def send_approval_request(self, workflow_id: str, stage_gate: "StageGate",
                              agent_name: str, project_name: str,
                              content_summary: str, content_detail: str) -> bool:
        if not self.enabled:
            logger.warning(f"SMTP disabled. Approval for {workflow_id} not sent.")
            return False

        subject = (f"[Action Required] {stage_gate.description} — "
                   f"{project_name} ({agent_name})")
        body = (
            f"APPROVAL REQUEST\nAgent: {agent_name}\nProject: {project_name}\n"
            f"Stage Gate: {stage_gate.description}\nWorkflow ID: {workflow_id}\n\n"
            f"SUMMARY:\n{content_summary}\n\nCONTENT:\n{content_detail[:500]}...\n\n"
            f"Reply APPROVED or REJECTED (with optional feedback).\n"
            f"Auto-escalates in {stage_gate.timeout_hours}h.\n"
        )
        try:
            msg = MIMEMultipart()
            msg["From"]    = self.sender_email
            msg["To"]      = stage_gate.approver_email
            msg["Subject"] = subject
            if stage_gate.cc_emails:
                msg["Cc"] = ", ".join(stage_gate.cc_emails)
            msg.attach(MIMEText(body, "plain"))

            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                server.starttls()
                server.login(self.sender_email, self.sender_password)
                recipients = [stage_gate.approver_email] + (stage_gate.cc_emails or [])
                server.sendmail(self.sender_email, recipients, msg.as_string())

            logger.info(f"SMTP email sent for {workflow_id} → {stage_gate.approver_email}")
            return True
        except Exception as e:
            logger.error(f"SMTP email failed for {workflow_id}: {e}")
            return False

    def parse_approval_response(self, email_subject: str, email_body: str) -> Tuple[str, str]:
        body_upper = email_body.upper()
        if "APPROVED" in body_upper:
            decision = "approved"
        elif "REJECTED" in body_upper:
            decision = "rejected"
        else:
            decision = "clarification_needed"
        return decision, email_body.strip()


# Keep EmailGateway as the legacy name so nothing else breaks
EmailGateway = SmtpEmailGateway


def create_email_gateway():
    """Factory — returns the right gateway based on EMAIL_PROVIDER env var.

    EMAIL_PROVIDER=graph  → GraphEmailGateway  (Microsoft Graph API, recommended)
    EMAIL_PROVIDER=smtp   → SmtpEmailGateway   (Outlook SMTP, legacy)
    EMAIL_PROVIDER=none   → SmtpEmailGateway   (disabled, no-op)
    default               → tries Graph first, falls back to SMTP
    """
    provider = os.environ.get("EMAIL_PROVIDER", "auto").lower()

    if provider == "graph":
        return GraphEmailGateway()
    elif provider == "smtp":
        return SmtpEmailGateway()
    elif provider == "none":
        logger.info("Email provider set to none — all gates will log only.")
        return SmtpEmailGateway()   # credentials missing → self.enabled=False → no-op
    else:
        # Auto: prefer Graph if configured, fall back to SMTP
        gw = GraphEmailGateway()
        if gw.enabled:
            logger.info("Email provider: Graph API (auto-selected)")
            return gw
        gw = SmtpEmailGateway()
        if gw.enabled:
            logger.info("Email provider: SMTP (auto-selected fallback)")
            return gw
        logger.warning("Email provider: none configured — gates will log only.")
        return gw   # returns disabled SMTP gateway as a safe no-op


# ============================================================================
class StageGateManager:
    """Manage stage gates and approval workflows."""

    STAGE_GATES = {
        StageGateName.CHARTER_APPROVAL: StageGate(
            name=StageGateName.CHARTER_APPROVAL,
            description="Project Charter Review & Approval",
            approver_email="tjohnson@firstgenesis.com",
            cc_emails=["pwatty@firstgenesis.com", "emathieu@firstgenesis.com"],
            require_comment=False,
            timeout_hours=24
        ),
        StageGateName.REQUIREMENTS_APPROVAL: StageGate(
            name=StageGateName.REQUIREMENTS_APPROVAL,
            description="Requirements Specification Review",
            approver_email="kphipps@firstgenesis.com",
            cc_emails=["tjohnson@firstgenesis.com", "emathieu@firstgenesis.com"],
            require_comment=True,
            timeout_hours=12
        ),
        StageGateName.QA_AUDIT_APPROVAL: StageGate(
            name=StageGateName.QA_AUDIT_APPROVAL,
            description="QA Audit & Pre-Delivery Approval",
            approver_email="emathieu@firstgenesis.com",
            cc_emails=["tjohnson@firstgenesis.com"],
            require_comment=False,
            timeout_hours=6
        ),
        StageGateName.DELIVERY_APPROVAL: StageGate(
            name=StageGateName.DELIVERY_APPROVAL,
            description="Final Approval Before Customer Delivery",
            approver_email="tjohnson@firstgenesis.com",
            cc_emails=["pwatty@firstgenesis.com"],
            require_comment=True,
            timeout_hours=2
        ),
        StageGateName.BUDGET_ESCALATION: StageGate(
            name=StageGateName.BUDGET_ESCALATION,
            description="Budget Alert - Approval to Continue",
            approver_email="pwatty@firstgenesis.com",
            cc_emails=["emathieu@firstgenesis.com"],
            require_comment=True,
            timeout_hours=1
        ),
        StageGateName.DOCUMENT_CLEARANCE: StageGate(
            name=StageGateName.DOCUMENT_CLEARANCE,
            description="Inbound Document PII Review Before LLM Ingestion",
            approver_email="kphipps@firstgenesis.com",
            cc_emails=["emathieu@firstgenesis.com", "tjohnson@firstgenesis.com"],
            require_comment=True,
            timeout_hours=4
        ),
        StageGateName.LAB_SIGN_OFF: StageGate(
            name=StageGateName.LAB_SIGN_OFF,
            description="Lab Validation Sign-Off Before Production Promotion",
            approver_email="tjohnson@firstgenesis.com",
            cc_emails=["pwatty@firstgenesis.com"],
            require_comment=True,
            timeout_hours=24
        ),
        StageGateName.SME_REVIEW: StageGate(
            name=StageGateName.SME_REVIEW,
            description="Low-Confidence Agent Output — SME Review Required",
            approver_email="kphipps@firstgenesis.com",
            cc_emails=["emathieu@firstgenesis.com"],
            require_comment=True,
            timeout_hours=8
        ),
        StageGateName.COMBINED_REVIEW: StageGate(
            name=StageGateName.COMBINED_REVIEW,
            description="Phase 1 — Agent Review Complete, Awaiting SME Approval",
            approver_email="kphipps@firstgenesis.com",
            cc_emails=["tjohnson@firstgenesis.com", "emathieu@firstgenesis.com",
                       "pwatty@firstgenesis.com"],
            require_comment=True,
            timeout_hours=24
        ),
    }

    def __init__(self, db: WorkflowDatabase, email_gateway: EmailGateway):
        self.db = db
        self.email = email_gateway

    def pause_at_gate(self, workflow_id: str, agent_name: str, project_name: str,
                      stage_gate_name: StageGateName,
                      content_pending_approval: str) -> WorkflowState:
        stage_gate = self.STAGE_GATES.get(stage_gate_name)
        if not stage_gate:
            raise ValueError(f"Unknown stage gate: {stage_gate_name}")

        content_hash = hashlib.sha256(content_pending_approval.encode()).hexdigest()
        workflow_state = WorkflowState(
            workflow_id=workflow_id, agent_name=agent_name, project_name=project_name,
            current_stage_gate=stage_gate_name, status=WorkflowStatus.PENDING,
            content_pending_approval=content_pending_approval, content_hash=content_hash
        )
        self.db.save_workflow_state(workflow_state)

        email_sent = self.email.send_approval_request(
            workflow_id=workflow_id, stage_gate=stage_gate,
            agent_name=agent_name, project_name=project_name,
            content_summary=content_pending_approval[:200],
            content_detail=content_pending_approval
        )

        if email_sent:
            workflow_state.status = WorkflowStatus.APPROVAL_SENT
            workflow_state.approval_email_sent_at = datetime.now().isoformat()
        else:
            logger.warning(f"Email failed for {workflow_id}, workflow remains in PENDING")

        self.db.save_workflow_state(workflow_state)
        return workflow_state

    def record_approval_response(self, workflow_id: str, approver_email: str,
                                 decision: str, feedback: str = None) -> WorkflowState:
        workflow_state = self.db.get_workflow_state(workflow_id)
        if not workflow_state:
            raise ValueError(f"Workflow not found: {workflow_id}")

        self.db.record_approval(
            workflow_id=workflow_id, stage_gate=workflow_state.current_stage_gate,
            approver_email=approver_email, decision=decision, feedback=feedback
        )

        if decision.lower() == "approved":
            workflow_state.status = WorkflowStatus.APPROVED
            workflow_state.human_approver = approver_email
            workflow_state.human_feedback = feedback
            workflow_state.approval_timestamp = datetime.now().isoformat()
            logger.info(f"Workflow {workflow_id} APPROVED by {approver_email}")
        elif decision.lower() == "rejected":
            workflow_state.status = WorkflowStatus.REJECTED
            workflow_state.human_approver = approver_email
            workflow_state.human_feedback = feedback
            logger.warning(f"Workflow {workflow_id} REJECTED by {approver_email}: {feedback}")
        else:
            logger.warning(f"Unknown decision for {workflow_id}: {decision}")

        self.db.save_workflow_state(workflow_state)
        return workflow_state

    def get_pending_approvals_summary(self) -> str:
        pending = self.db.get_pending_approvals()
        if not pending:
            return "No pending approvals"

        summary = f"\n{'='*70}\nPENDING APPROVALS ({len(pending)})\n{'='*70}\n"
        for wf in pending:
            sent_at = datetime.fromisoformat(wf.approval_email_sent_at)
            hours_waiting = (datetime.now() - sent_at).total_seconds() / 3600
            summary += (f"\n[{wf.workflow_id}]\n"
                        f"  Agent:      {wf.agent_name}\n"
                        f"  Project:    {wf.project_name}\n"
                        f"  Stage Gate: {wf.current_stage_gate.value}\n"
                        f"  Waiting:    {hours_waiting:.1f} hours\n"
                        f"  Content:    {wf.content_pending_approval[:80]}...\n")
        return summary + "\n" + "=" * 70

# ============================================================================
# BUDGET ENFORCER
# ============================================================================
class BudgetEnforcer:
    """Enforces daily and per-agent budget limits."""

    DAILY_BUDGET_USD = 5.00
    ALERT_THRESHOLD = 0.80

    AGENT_CONFIGS = {
        "pm_agent": AgentConfig(
            name="pm_agent", budget_per_call_usd=0.03, max_daily_calls=2,
            max_daily_spend_usd=0.10, priority=1, description="Project setup, WBS, status"
        ),
        "ba_agent": AgentConfig(
            name="ba_agent", budget_per_call_usd=0.04, max_daily_calls=5,
            max_daily_spend_usd=0.25, priority=1,
            description="Requirements, design sessions, data intake, schema extraction"
        ),
        "qa_agent": AgentConfig(
            name="qa_agent", budget_per_call_usd=0.06, max_daily_calls=1,
            max_daily_spend_usd=0.10, priority=2, description="Pre-delivery audit"
        ),
        "vendor_agent": AgentConfig(
            name="vendor_agent", budget_per_call_usd=0.02, max_daily_calls=1,
            max_daily_spend_usd=0.05, priority=3, description="Partner SLA tracking"
        ),
        "manager_agent": AgentConfig(
            name="manager_agent", budget_per_call_usd=0.04, max_daily_calls=1,
            max_daily_spend_usd=0.10, priority=1, description="Portfolio dashboard"
        ),
    }

    def __init__(self, db: WorkflowDatabase):
        self.db = db

    def can_call_agent(self, agent_name: str, estimated_cost: float) -> Tuple[bool, str]:
        if agent_name not in self.AGENT_CONFIGS:
            return False, f"Unknown agent: {agent_name}"

        config = self.AGENT_CONFIGS[agent_name]
        today_spend, _ = self.db.get_today_spend()
        agent_spend, agent_calls = self.db.get_agent_spend_today(agent_name)

        if today_spend + estimated_cost > self.DAILY_BUDGET_USD:
            msg = (f"Daily budget exceeded: ${today_spend:.2f} + ${estimated_cost:.2f} "
                   f"> ${self.DAILY_BUDGET_USD:.2f}")
            logger.error(msg)
            return False, msg

        if agent_spend + estimated_cost > config.max_daily_spend_usd:
            msg = (f"{agent_name} daily limit exceeded: ${agent_spend:.2f} + "
                   f"${estimated_cost:.2f} > ${config.max_daily_spend_usd:.2f}")
            logger.error(msg)
            return False, msg

        if agent_calls + 1 > config.max_daily_calls:
            msg = f"{agent_name} call limit: {agent_calls} + 1 > {config.max_daily_calls}"
            logger.error(msg)
            return False, msg

        projected = today_spend + estimated_cost
        if projected > self.DAILY_BUDGET_USD * self.ALERT_THRESHOLD:
            logger.warning(f"Budget alert: {agent_name} will push spend to "
                           f"${projected:.2f} ({projected/self.DAILY_BUDGET_USD*100:.0f}%)")

        return True, "Budget OK"

    def get_status_report(self) -> str:
        today_spend, today_calls = self.db.get_today_spend()
        week_spend = self.db.get_last_n_days_spend(7)

        report = (f"\n{'='*70}\n"
                  f"BUDGET STATUS REPORT - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                  f"{'='*70}\n"
                  f"Today's Spend:        ${today_spend:.4f} / ${self.DAILY_BUDGET_USD:.2f} "
                  f"({today_spend/self.DAILY_BUDGET_USD*100:.1f}%)\n"
                  f"Today's Calls:        {today_calls}\n"
                  f"Last 7 Days:          ${week_spend:.2f}\n"
                  f"Daily Average (7d):   ${week_spend/7:.2f}\n"
                  f"Per-Agent Today:\n")

        for agent_name, config in self.AGENT_CONFIGS.items():
            agent_spend, agent_calls = self.db.get_agent_spend_today(agent_name)
            pct = (agent_spend / config.max_daily_spend_usd * 100) if config.max_daily_spend_usd else 0
            report += (f"  {agent_name:20} ${agent_spend:.4f} / ${config.max_daily_spend_usd:.2f} "
                       f"({pct:3.0f}%) [{agent_calls} calls]\n")

        return report + "\n" + "=" * 70

# ============================================================================
# HALLUCINATION GUARD
# ============================================================================
class HallucinationGuard:
    """Detect and prevent agent hallucinations.

    Frozen facts can be updated at runtime by SMEs via add_frozen_fact().
    Updates persist in the workflow database and are loaded on init.
    """

    # Hardcoded baseline — these are always active
    _BASE_FROZEN_FACTS = {
        "project_timeline_aura": "3 months",
        "project_client_aura": "Malcolm Goodwin",
        "daily_budget": "$5 USD",
        "team_pm": "Kiera Phipps",
        "team_cto": "Ron Watty",
        "team_cdo": "Trice Johnson",
        "team_pmo": "Elina Mathieu",
        "team_ceo": "Pascal Watty",
        "vendor_wbt": "Yubi",
        "deadline_wbt": "April 30, 2026"
    }

    IMPOSSIBLE_CLAIMS = [
        (r"Chevron.*approved", "Chevron not approved yet"),
        (r"AURA.*complete|AURA.*finished", "AURA just kicked off"),
        (r"new hire|brought on|onboarded", "No hiring decisions made yet"),
        (r"first genesis.*failed|bankruptcy|shutdown", "Company operational"),
    ]

    def __init__(self, db: WorkflowDatabase):
        self.db = db
        # Merge base facts with any SME additions stored in DB
        self.FROZEN_FACTS = dict(self._BASE_FROZEN_FACTS)
        self._load_dynamic_facts()

    def _load_dynamic_facts(self):
        """Load SME-added facts from the database."""
        try:
            self.db.cursor.execute(
                "SELECT fact_key, fact_value FROM frozen_facts WHERE active = 1"
            )
            for key, value in self.db.cursor.fetchall():
                self.FROZEN_FACTS[key] = value
        except sqlite3.OperationalError:
            pass  # Table may not exist yet on first run

    def add_frozen_fact(self, key: str, value: str, added_by: str = "sme"):
        """SMEs can add new ground-truth facts at runtime."""
        self.FROZEN_FACTS[key] = value
        try:
            self.db.cursor.execute("""
                INSERT OR REPLACE INTO frozen_facts (fact_key, fact_value, added_by, created_at, active)
                VALUES (?, ?, ?, ?, 1)
            """, (key, value, added_by, datetime.now().isoformat()))
            self.db.conn.commit()
            logger.info(f"Frozen fact added by {added_by}: {key} = {value}")
        except Exception as e:
            logger.error(f"Failed to persist frozen fact: {e}")

    def remove_frozen_fact(self, key: str):
        """Deactivate a dynamic frozen fact (base facts cannot be removed)."""
        if key in self._BASE_FROZEN_FACTS:
            logger.warning(f"Cannot remove base frozen fact: {key}")
            return
        self.FROZEN_FACTS.pop(key, None)
        try:
            self.db.cursor.execute(
                "UPDATE frozen_facts SET active = 0 WHERE fact_key = ?", (key,)
            )
            self.db.conn.commit()
        except Exception as e:
            logger.error(f"Failed to deactivate frozen fact: {e}")

    def validate_output(self, agent_name: str, output: str) -> Tuple[bool, str]:
        output_lower = output.lower()

        # Check frozen fact contradictions
        if "aura" in agent_name or "aura" in output_lower:
            if ("4 month" in output_lower or "6 month" in output_lower) and \
               "3 month" not in output_lower:
                reason = "Agent timeline contradicts frozen fact (3 months)"
                self.db.log_hallucination_flag(agent_name, reason, output)
                return False, reason

        # Check impossible claims
        for pattern, reason_text in self.IMPOSSIBLE_CLAIMS:
            if re.search(pattern, output, re.IGNORECASE):
                reason = f"Impossible claim detected: {reason_text}"
                self.db.log_hallucination_flag(agent_name, reason, output)
                return False, reason

        # Soft warning for undocumented project claims
        if "project" in output_lower and agent_name in ("pm_agent", "ba_agent"):
            if "document" not in output_lower and "template" not in output_lower:
                self.db.log_hallucination_flag(
                    agent_name, "Warning: project claim without doc reference", output
                )

        return True, "Hallucination check passed"

# ============================================================================
# PII REDACTOR
# ============================================================================
class PIIRedactor:
    """Scrubs PII from inbound documents before any LLM touches them.

    Uses regex patterns to replace sensitive values with typed tokens.
    Stores a reverse map (token -> original) in SQLite so the original
    values can be restored after human review if needed.

    Usage:
        redactor = PIIRedactor()
        redacted_text, reverse_map = redactor.redact(raw_text)
        # ... send redacted_text to LLM ...
        original_text = redactor.restore(llm_output, reverse_map)
    """

    # (pattern, placeholder_prefix) — order matters: more specific first
    PATTERNS = [
        (r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', 'EMAIL'),
        (r'\b\d{3}[-.\s]?\d{3}[-.\s]?\d{4}\b', 'PHONE'),
        (r'\$[\d,]+(?:\.\d{2})?|\b\d{1,3}(?:,\d{3})+(?:\.\d{2})?\b', 'AMOUNT'),
        (r'\b\d{3}-\d{2}-\d{4}\b', 'SSN'),
        (r'\b(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|'
         r'Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|'
         r'Dec(?:ember)?)[.\s]+\d{1,2}[,.\s]+\d{4}\b', 'DATE'),
        (r'\b\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}\b', 'DATE'),
    ]

    def __init__(self):
        self._counter = 0

    def redact(self, text: str) -> Tuple[str, Dict[str, str]]:
        """Replace PII with tokens. Returns (redacted_text, reverse_map)."""
        reverse_map: Dict[str, str] = {}
        redacted = text
        for pattern, prefix in self.PATTERNS:
            for match in re.finditer(pattern, redacted, re.IGNORECASE):
                original = match.group(0)
                # Reuse token if same value already seen
                existing = next(
                    (tok for tok, val in reverse_map.items() if val == original), None
                )
                if existing:
                    continue
                token = f"[{prefix}_{self._counter:04d}]"
                self._counter += 1
                reverse_map[token] = original
                redacted = redacted.replace(original, token)
        return redacted, reverse_map

    @staticmethod
    def restore(text: str, reverse_map: Dict[str, str]) -> str:
        """Re-insert original values using the reverse map."""
        for token, original in reverse_map.items():
            text = text.replace(token, original)
        return text

    def redact_dict(self, data: dict) -> Tuple[dict, Dict[str, str]]:
        """Recursively redact all string values in a dict."""
        combined_map: Dict[str, str] = {}
        result = {}
        for key, value in data.items():
            if isinstance(value, str):
                redacted, rmap = self.redact(value)
                result[key] = redacted
                combined_map.update(rmap)
            elif isinstance(value, dict):
                redacted_sub, rmap = self.redact_dict(value)
                result[key] = redacted_sub
                combined_map.update(rmap)
            elif isinstance(value, list):
                new_list = []
                for item in value:
                    if isinstance(item, str):
                        r, rmap = self.redact(item)
                        new_list.append(r)
                        combined_map.update(rmap)
                    else:
                        new_list.append(item)
                result[key] = new_list
            else:
                result[key] = value
        return result, combined_map


# ============================================================================
# AUTONOMOUS AGENT WITH EMAIL GATES + GUARDRAILS
# ============================================================================
class AutonomousAgentWithEmailGates:
    """Agent system with email approval gates, budget enforcement, and hallucination detection."""

    def __init__(self, api_key: Optional[str] = None):
        self.client = anthropic.Anthropic(api_key=api_key or os.environ.get("ANTHROPIC_API_KEY"))
        self.db = WorkflowDatabase()
        self.email_gateway = create_email_gateway()
        self.stage_gate_manager = StageGateManager(self.db, self.email_gateway)
        self.budget_enforcer = BudgetEnforcer(self.db)
        self.hallucination_guard = HallucinationGuard(self.db)
        self.knowledge_library = KnowledgeLibrary()
        self.pii_redactor = PIIRedactor()
        logger.info(f"Agent system initialized (env={FG_ENV}, model={ACTIVE_MODEL})")

    def _build_system_prompt_with_corrections(self, agent_name: str,
                                               base_system_prompt: str) -> str:
        """Inject top SME corrections as few-shot examples into the system prompt."""
        try:
            corrections = self.knowledge_library.get_sme_corrections(agent_name, limit=3)
        except Exception:
            corrections = []
        if not corrections:
            return base_system_prompt
        correction_block = "\n\nSME CORRECTIONS (learn from these past mistakes):\n"
        for i, c in enumerate(corrections, 1):
            correction_block += (
                f"\nExample {i} — {c['category']}:\n"
                f"  Wrong output:   {c['original_snippet'][:150]}\n"
                f"  Correct output: {c['corrected_content'][:150]}\n"
            )
        return base_system_prompt + correction_block

    def _call_claude(self, agent_name: str, system_prompt: str,
                     user_message: str, workflow_id: str = None,
                     scaffold_mode: bool = True) -> Tuple[str, AgentCall, int]:
        """Call Claude with budget check, hallucination validation, and confidence scoring.

        Args:
            scaffold_mode: When True, instructs agent to populate templates only
                           and never author design sections marked HUMAN_ONLY.
        Returns:
            (output_text, AgentCall, confidence_score 1-5)
        """
        # In lab mode, relax budget checks
        budget_multiplier = LAB_BUDGET_MULTIPLIER

        # Estimate cost (use actual model pricing)
        estimated_tokens = len(user_message.split()) * 1.3
        input_cost_rate = 0.25 if IS_LAB_MODE else TokenCost.INPUT_COST_PER_1M
        estimated_cost = (estimated_tokens / 1_000_000) * input_cost_rate

        # Budget check
        can_proceed, budget_reason = self.budget_enforcer.can_call_agent(
            agent_name, estimated_cost / budget_multiplier
        )
        if not can_proceed and not IS_LAB_MODE:
            call = AgentCall(
                agent_name=agent_name, timestamp=datetime.now().isoformat(),
                input_tokens=0, output_tokens=0, cache_created=0, cache_read=0,
                cost_usd=0, status="rejected_budget", reason=budget_reason, output_hash=""
            )
            self.db.log_agent_call(call, workflow_id)
            logger.error(f"{agent_name}: Budget rejected - {budget_reason}")
            return "", call, 0

        # Inject SME corrections into system prompt
        enriched_system = self._build_system_prompt_with_corrections(agent_name, system_prompt)

        # Scaffold mode: prevent agents from authoring HUMAN_ONLY sections
        if scaffold_mode:
            enriched_system += (
                "\n\nSCAFFOLD MODE — CRITICAL RULES:\n"
                "1. Populate template fields with EXTRACTED FACTS only.\n"
                "2. Any field marked [HUMAN_ONLY] must be left blank — output empty string.\n"
                "3. Add a 'confidence' field (integer 1-5) to your JSON output.\n"
                "   1=guessing, 3=reasonable, 5=certain from source data.\n"
                "4. If you are not certain of a value, output [NEEDS_SME_INPUT] instead.\n"
                "5. Tag each generated sentence with [AGENT_DRAFTED] or [FACT_EXTRACTED].\n"
            )
        else:
            enriched_system += "\n\nAlways include a 'confidence' field (1-5) in your JSON output."

        # API call — use ACTIVE_MODEL based on environment
        try:
            response = self.client.messages.create(
                model=ACTIVE_MODEL,
                max_tokens=2000,
                system=enriched_system,
                messages=[{"role": "user", "content": user_message}]
            )
        except Exception as e:
            call = AgentCall(
                agent_name=agent_name, timestamp=datetime.now().isoformat(),
                input_tokens=0, output_tokens=0, cache_created=0, cache_read=0,
                cost_usd=0, status="error", reason=str(e), output_hash=""
            )
            self.db.log_agent_call(call, workflow_id)
            logger.error(f"{agent_name}: API error - {str(e)}")
            return "", call, 0

        output = response.content[0].text
        output_hash = hashlib.sha256(output.encode()).hexdigest()
        cost = TokenCost(
            input_tokens=response.usage.input_tokens,
            output_tokens=response.usage.output_tokens,
            cache_creation_tokens=getattr(response.usage, 'cache_creation_input_tokens', 0),
            cache_read_tokens=getattr(response.usage, 'cache_read_input_tokens', 0)
        ).total_cost_usd()

        # Extract confidence score from JSON output
        confidence = 5  # default high if no score provided
        try:
            parsed = json.loads(output)
            confidence = int(parsed.get("confidence", 5))
            confidence = max(1, min(5, confidence))
        except (json.JSONDecodeError, ValueError, TypeError):
            # Non-JSON response — attempt regex extraction
            m = re.search(r'"confidence"\s*:\s*(\d)', output)
            if m:
                confidence = max(1, min(5, int(m.group(1))))

        # Hallucination check
        is_valid, hall_reason = self.hallucination_guard.validate_output(agent_name, output)
        if not is_valid:
            call = AgentCall(
                agent_name=agent_name, timestamp=datetime.now().isoformat(),
                input_tokens=response.usage.input_tokens,
                output_tokens=response.usage.output_tokens,
                cache_created=getattr(response.usage, 'cache_creation_input_tokens', 0),
                cache_read=getattr(response.usage, 'cache_read_input_tokens', 0),
                cost_usd=cost, status="rejected_hallucination",
                reason=hall_reason, output_hash=output_hash
            )
            self.db.log_agent_call(call, workflow_id)
            logger.error(f"{agent_name}: Hallucination detected - {hall_reason}")
            return "", call, 0

        call = AgentCall(
            agent_name=agent_name, timestamp=datetime.now().isoformat(),
            input_tokens=response.usage.input_tokens,
            output_tokens=response.usage.output_tokens,
            cache_created=getattr(response.usage, 'cache_creation_input_tokens', 0),
            cache_read=getattr(response.usage, 'cache_read_input_tokens', 0),
            cost_usd=cost, status="success", reason="", output_hash=output_hash
        )
        self.db.log_agent_call(call, workflow_id)
        logger.info(f"{agent_name}: Success (confidence={confidence}) - {call}")
        return output, call, confidence

    def run_pm_agent_with_gates(self, project_metadata: dict) -> Tuple[str, WorkflowState]:
        """Run PM Agent with guardrails + charter approval gate.

        Uses scaffold mode: agent fills in known facts only.
        Low-confidence output (<3) automatically routes to SME_REVIEW gate
        before proceeding to CHARTER_APPROVAL.
        """
        workflow_id = f"pm_{project_metadata.get('project','proj')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        project_name = project_metadata.get('project', 'Unknown')
        logger.info(f"Starting PM Agent workflow: {workflow_id}")

        # In lab mode, mark workflow for lab sign-off before production promotion
        env_tag = "[LAB]" if IS_LAB_MODE else "[PROD]"

        output, call, confidence = self._call_claude(
            agent_name="pm_agent",
            system_prompt=(
                "You are a Project Manager Agent for First Genesis. "
                "Output ONLY valid JSON. NO explanations, NO preamble. "
                "You ASSIST human architects — do NOT make design decisions."
            ),
            user_message=f"""Populate the project charter template with EXTRACTED FACTS only.
Project metadata:
{json.dumps(project_metadata, indent=2)}

Output JSON only — fill each field from the source data above.
Mark any field you cannot determine from the data as "[NEEDS_SME_INPUT]".
Sections marked [HUMAN_ONLY] must be empty strings.

{{"project_charter": {{
  "title": "...",
  "client": "...",
  "timeline": "3 months",
  "scope": "...",
  "success_criteria": "[HUMAN_ONLY]",
  "technical_architecture": "[HUMAN_ONLY]",
  "integration_strategy": "[HUMAN_ONLY]"
}},
"wbs": {{}},
"risks": [],
"confidence": 4,
"env": "{env_tag}"
}}""",
            workflow_id=workflow_id,
            scaffold_mode=True
        )

        if not output:
            raise RuntimeError(f"PM Agent call failed: {call.reason}")

        # Confidence < 3: route to SME review before charter approval
        if confidence < 3:
            logger.warning(
                f"PM Agent low confidence ({confidence}/5) — routing to SME_REVIEW gate"
            )
            workflow_state = self.stage_gate_manager.pause_at_gate(
                workflow_id=workflow_id, agent_name="pm_agent", project_name=project_name,
                stage_gate_name=StageGateName.SME_REVIEW,
                content_pending_approval=f"[confidence={confidence}/5]\n{output}"
            )
        elif IS_LAB_MODE:
            # Lab workflows go to LAB_SIGN_OFF before charter approval
            workflow_state = self.stage_gate_manager.pause_at_gate(
                workflow_id=workflow_id, agent_name="pm_agent", project_name=project_name,
                stage_gate_name=StageGateName.LAB_SIGN_OFF,
                content_pending_approval=output
            )
        else:
            workflow_state = self.stage_gate_manager.pause_at_gate(
                workflow_id=workflow_id, agent_name="pm_agent", project_name=project_name,
                stage_gate_name=StageGateName.CHARTER_APPROVAL,
                content_pending_approval=output
            )

        logger.info(
            f"Workflow {workflow_id} paused at {workflow_state.current_stage_gate.value} "
            f"(confidence={confidence}/5)"
        )
        return output, workflow_state

    def resume_approved_workflow(self, workflow_id: str) -> str:
        workflow_state = self.db.get_workflow_state(workflow_id)
        if not workflow_state:
            logger.error(f"Workflow not found: {workflow_id}")
            return ""

        if workflow_state.status != WorkflowStatus.APPROVED:
            logger.warning(f"Workflow {workflow_id} not APPROVED: {workflow_state.status.value}")
            return ""

        next_actions = {
            StageGateName.CHARTER_APPROVAL: "Charter approved. Ready for customer kickoff.",
            StageGateName.REQUIREMENTS_APPROVAL: "Requirements approved. Starting design sessions.",
            StageGateName.QA_AUDIT_APPROVAL: "Audit approved. Proceeding to delivery.",
        }

        next_action = next_actions.get(workflow_state.current_stage_gate, "")
        if not next_action:
            logger.warning(f"Unknown next step for gate: {workflow_state.current_stage_gate.value}")
            return ""

        is_final = workflow_state.current_stage_gate == StageGateName.QA_AUDIT_APPROVAL
        workflow_state.status = WorkflowStatus.COMPLETED if is_final else WorkflowStatus.RESUMED
        workflow_state.next_step_after_approval = next_action
        self.db.save_workflow_state(workflow_state)
        logger.info(f"Agent proceeding: {next_action}")
        return next_action

    def check_pending_approvals(self) -> str:
        return self.stage_gate_manager.get_pending_approvals_summary()

    def process_approval_response(self, workflow_id: str, approver_email: str,
                                  decision: str, feedback: str = None) -> str:
        workflow_state = self.stage_gate_manager.record_approval_response(
            workflow_id=workflow_id, approver_email=approver_email,
            decision=decision, feedback=feedback
        )
        return (f"\nApproval recorded for {workflow_id}:\n"
                f"  Decision:  {decision}\n"
                f"  Approver:  {approver_email}\n"
                f"  Feedback:  {feedback or 'None'}\n\n"
                f"Workflow status updated to: {workflow_state.status.value}\n")

    def run_ba_agent_requirements(self, project_metadata: dict) -> Tuple[str, WorkflowState]:
        """Run BA Agent for requirements extraction — routes to REQUIREMENTS_APPROVAL gate."""
        workflow_id = f"ba_req_{project_metadata.get('project','proj')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        project_name = project_metadata.get('project', 'Unknown')
        logger.info(f"Starting BA Agent requirements workflow: {workflow_id}")

        output, call, confidence = self._call_claude(
            agent_name="ba_agent",
            system_prompt=(
                "You are a Business Analyst Agent for First Genesis. "
                "Extract and document requirements from provided project data. "
                "Output ONLY valid JSON. You ASSIST human architects — do NOT make design decisions. "
                "Sections marked [HUMAN_ONLY] must be left as empty strings."
            ),
            user_message=f"""Extract requirements from this project data.
Project: {json.dumps(project_metadata, indent=2)}

Populate the requirements template with EXTRACTED FACTS only.
Mark unknown fields as \"[NEEDS_SME_INPUT]\".

{{"requirements": {{
  "functional": [
    {{"id": "FR1", "description": "...", "priority": "high|medium|low",
      "acceptance_criteria": "...", "source": "[FACT_EXTRACTED] or [NEEDS_SME_INPUT]"}}
  ],
  "non_functional": [
    {{"id": "NFR1", "category": "performance|security|usability", "description": "..."}}
  ],
  "assumptions": [],
  "constraints": [],
  "integration_design": "[HUMAN_ONLY]",
  "technical_approach": "[HUMAN_ONLY]"
}},
"traceability_matrix": [],
"confidence": 4
}}""",
            workflow_id=workflow_id,
            scaffold_mode=True
        )

        if not output:
            raise RuntimeError(f"BA Agent call failed: {call.reason}")

        gate = StageGateName.SME_REVIEW if confidence < 3 else StageGateName.REQUIREMENTS_APPROVAL
        workflow_state = self.stage_gate_manager.pause_at_gate(
            workflow_id=workflow_id, agent_name="ba_agent", project_name=project_name,
            stage_gate_name=gate,
            content_pending_approval=f"[confidence={confidence}/5]\n{output}"
        )
        logger.info(f"BA requirements workflow {workflow_id} paused at {gate.value} (confidence={confidence}/5)")
        return output, workflow_state

    def run_ba_agent_bdr_intake(self, filepath: str) -> Tuple[str, WorkflowState]:
        """BA Agent ingests a BDR document (CSV or Excel).

        Pipeline:
        1. Parse file into rows (CSV stdlib / xlsx via openpyxl)
        2. PII redaction on all string fields
        3. Pause at DOCUMENT_CLEARANCE gate for human review
        After approval, call complete_ba_data_analysis() to extract schema + requirements.
        """
        doc_id = f"bdr_{uuid.uuid4().hex[:12]}"
        workflow_id = f"ba_intake_{doc_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        filename = os.path.basename(filepath)
        file_ext = os.path.splitext(filename)[1].lower()
        logger.info(f"BA Agent BDR intake started: {filename} (doc_id={doc_id})")

        raw_rows: List[Dict] = []
        try:
            if file_ext == ".csv":
                with open(filepath, "r", encoding="utf-8-sig") as f:
                    raw_rows = [dict(row) for row in csv.DictReader(f)]
            elif file_ext in (".xlsx", ".xls"):
                if not OPENPYXL_AVAILABLE:
                    raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    raw_rows.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
                wb.close()
            else:
                raise ValueError(f"Unsupported file type: {file_ext}. Supported: .csv, .xlsx")
        except Exception as e:
            logger.error(f"BA Agent BDR parse failed for {filename}: {e}")
            raise

        logger.info(f"Parsed {len(raw_rows)} rows from {filename}")

        # PII redaction
        redacted_rows, combined_map = [], {}
        for row in raw_rows:
            r, rmap = self.pii_redactor.redact_dict(row)
            redacted_rows.append(r)
            combined_map.update(rmap)
        logger.info(f"PII redaction: {len(combined_map)} tokens replaced across {len(raw_rows)} rows")

        # Persist pre-clearance record
        self.db.cursor.execute("""
            INSERT OR REPLACE INTO bdr_documents
            (doc_id, filename, file_type, raw_row_count, extracted_fields,
             redaction_map, clearance_workflow_id, status, ingested_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'pending_clearance', ?)
        """, (
            doc_id, filename, file_ext, len(raw_rows),
            json.dumps(redacted_rows[:5]),
            json.dumps(combined_map),
            workflow_id, datetime.now().isoformat()
        ))
        self.db.conn.commit()

        clearance_content = (
            f"BDR Document: {filename}\n"
            f"Rows: {len(raw_rows)} | PII tokens redacted: {len(combined_map)}\n"
            f"Sample (first row, redacted):\n{json.dumps(redacted_rows[0] if redacted_rows else {}, indent=2)}\n\n"
            f"APPROVE to allow BA Agent to extract schema and requirements.\n"
            f"REJECT to hold document — no LLM contact."
        )
        workflow_state = self.stage_gate_manager.pause_at_gate(
            workflow_id=workflow_id, agent_name="ba_agent", project_name=filename,
            stage_gate_name=StageGateName.DOCUMENT_CLEARANCE,
            content_pending_approval=clearance_content
        )
        logger.info(f"BA BDR intake paused at DOCUMENT_CLEARANCE (workflow_id={workflow_id})")

        return json.dumps({
            "doc_id": doc_id, "workflow_id": workflow_id, "filename": filename,
            "row_count": len(raw_rows), "pii_tokens_redacted": len(combined_map),
            "status": "pending_clearance",
            "next": f"After approval run: complete_ba_data_analysis --doc-id {doc_id}"
        }, indent=2), workflow_state

    def complete_ba_data_analysis(self, doc_id: str) -> str:
        """BA Agent extracts schema and requirements from a cleared BDR document."""
        self.db.cursor.execute("SELECT * FROM bdr_documents WHERE doc_id = ?", (doc_id,))
        row = self.db.cursor.fetchone()
        if not row:
            return f"ERROR: Document {doc_id} not found"

        _, filename, file_type, row_count, extracted_fields, \
            _, clearance_wf_id, status, ingested_at, _ = row

        if status == "pending_clearance":
            return f"ERROR: Document {doc_id} not yet cleared (still at DOCUMENT_CLEARANCE gate)."

        redacted_sample = json.loads(extracted_fields or "[]")
        workflow_id = f"ba_analysis_{doc_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        output, call, confidence = self._call_claude(
            agent_name="ba_agent",
            system_prompt=(
                "You are a Business Analyst Agent for First Genesis. "
                "Analyze business data documents to extract schemas, data requirements, "
                "and dependency maps. Output ONLY valid JSON. "
                "Sections marked [HUMAN_ONLY] must be empty strings."
            ),
            user_message=f"""Analyze this cleared BDR document and extract structured requirements.
Filename: {filename} | Rows: {row_count} | Type: {file_type}
Redacted sample: {json.dumps(redacted_sample, indent=2)}

Output JSON only:
{{
  "schema": {{"field_name": "inferred_type"}},
  "data_quality": {{"completeness_pct": 95, "issues": []}},
  "requirements_extracted": [
    {{"id": "DR1", "description": "...", "source_field": "..."}}
  ],
  "sme_review_fields": [],
  "dependency_map": {{
    "upstream": [], "downstream": [], "missing": []
  }},
  "integration_design": "[HUMAN_ONLY]",
  "technical_data_flow": "[HUMAN_ONLY]",
  "confidence": 4
}}""",
            workflow_id=workflow_id,
            scaffold_mode=True
        )

        if not output:
            return f"ERROR: BA Agent data analysis failed: {call.reason}"

        self.db.cursor.execute(
            "UPDATE bdr_documents SET status = 'analyzed', cleared_at = ? WHERE doc_id = ?",
            (datetime.now().isoformat(), doc_id)
        )
        self.db.conn.commit()
        logger.info(f"BA Agent completed data analysis for {filename} (confidence={confidence}/5)")
        return output

    def get_bdr_queue_status(self) -> str:
        """Return a summary of all BDR documents in the pipeline."""
        self.db.cursor.execute("""
            SELECT status, COUNT(*) as cnt, MAX(ingested_at) as latest
            FROM bdr_documents GROUP BY status
        """)
        rows = self.db.cursor.fetchall()
        lines = ["\n" + "="*70, "BA AGENT — BDR DOCUMENT QUEUE", "="*70]
        if not rows:
            lines.append("  No documents ingested yet.")
        for status, count, latest in rows:
            lines.append(f"  {status:30} {count:>3} doc(s)  latest: {latest}")

        # Flag documents stuck in clearance > 4 hours
        self.db.cursor.execute("""
            SELECT filename, ingested_at FROM bdr_documents WHERE status = 'pending_clearance'
        """)
        alerts = []
        for fname, ingested_at in self.db.cursor.fetchall():
            try:
                age_h = (datetime.now() - datetime.fromisoformat(ingested_at)).total_seconds() / 3600
                if age_h > 4:
                    alerts.append(f"  [STALE] '{fname}' waiting {age_h:.1f}h for clearance approval")
            except (ValueError, TypeError):
                pass
        if alerts:
            lines.append(f"\nALERTS ({len(alerts)}):")
            lines.extend(alerts)
        else:
            lines.append("\n  No stale documents.")
        lines.append("="*70 + "\n")
        return "\n".join(lines)

    # ── PHASE 1: VENDOR DELIVERABLE REVIEW ───────────────────────────────────
    # Agents do NOT create deliverables in Phase 1.
    # They intake vendor submissions, review them, and pass to SME for approval.

    def intake_vendor_deliverable(self, filepath: str, vendor_name: str,
                                   deliverable_type: str) -> str:
        """Phase 1 entry point — ingest a vendor-submitted deliverable.

        Pipeline:
        1. Parse file (text, CSV, xlsx, PDF-placeholder)
        2. PII redact content before any LLM ingestion
        3. Store in vendor_deliverables table (status=pending)
        4. DOCUMENT_CLEARANCE gate fires for human PII review
        5. After clearance, call run_vendor_deliverable_review(deliverable_id)
        """
        deliverable_id = f"del_{uuid.uuid4().hex[:12]}"
        filename = os.path.basename(filepath)
        ext = os.path.splitext(filename)[1].lower()
        now = datetime.now().isoformat()
        logger.info(f"Intake: {filename} from {vendor_name} (id={deliverable_id})")

        # ── Parse file content ───────────────────────────────────────────────
        raw_text = ""
        try:
            if ext == ".csv":
                with open(filepath, "r", encoding="utf-8-sig") as f:
                    raw_text = f.read()
            elif ext in (".xlsx", ".xls"):
                if not OPENPYXL_AVAILABLE:
                    raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                ws = wb.active
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append("\t".join([str(v) if v is not None else "" for v in row]))
                raw_text = "\n".join(rows)
                wb.close()
            else:
                # Treat as plain text / markdown / PDF stub
                with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                    raw_text = f.read()
        except Exception as e:
            logger.error(f"File parse error: {e}")
            return f"ERROR: Could not parse file '{filename}': {e}"

        # ── PII redact before storing ────────────────────────────────────────
        redacted_text, pii_map = self.pii_redactor.redact(raw_text)
        pii_count = len(pii_map)

        # ── Persist deliverable record ───────────────────────────────────────
        self.db.cursor.execute("""
            INSERT INTO vendor_deliverables
            (deliverable_id, vendor_name, deliverable_type, filename,
             content_text, submitted_at, review_status, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, 'pending', ?)
        """, (deliverable_id, vendor_name, deliverable_type, filename,
              redacted_text, now, now))
        self.db.conn.commit()

        # ── DOCUMENT_CLEARANCE gate (PII review before LLM) ─────────────────
        workflow_id = f"intake_{deliverable_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        clearance_summary = (
            f"VENDOR DELIVERABLE INTAKE — PII CLEARANCE REQUIRED\n"
            f"{'='*60}\n"
            f"  Deliverable ID:  {deliverable_id}\n"
            f"  Vendor:          {vendor_name}\n"
            f"  Type:            {deliverable_type}\n"
            f"  File:            {filename}\n"
            f"  PII Tokens Found:{pii_count}\n"
            f"{'='*60}\n"
            f"REDACTED PREVIEW (first 600 chars):\n"
            f"{redacted_text[:600]}\n"
            f"{'='*60}\n"
            f"ACTION: Review redacted content above.\n"
            f"If PII removal is satisfactory, reply APPROVED.\n"
            f"If additional redaction is needed, reply REJECTED with details.\n"
            f"After approval, the system will run agent review automatically.\n"
        )
        wf = self.stage_gate_manager.pause_at_gate(
            workflow_id=workflow_id,
            agent_name="ba_agent",
            project_name="AURA MVP",
            stage_gate_name=StageGateName.DOCUMENT_CLEARANCE,
            content_pending_approval=clearance_summary
        )
        # Link the clearance workflow back to the deliverable
        self.db.cursor.execute(
            "UPDATE vendor_deliverables SET workflow_id=?, review_status='pending_clearance', "
            "updated_at=? WHERE deliverable_id=?",
            (workflow_id, datetime.now().isoformat(), deliverable_id)
        )
        self.db.conn.commit()

        gate = StageGateManager.STAGE_GATES[StageGateName.DOCUMENT_CLEARANCE]
        return (
            f"\n{'='*70}\n"
            f"DELIVERABLE INTAKE COMPLETE — AWAITING PII CLEARANCE\n"
            f"{'='*70}\n"
            f"  Deliverable ID:  {deliverable_id}\n"
            f"  Vendor:          {vendor_name}\n"
            f"  Type:            {deliverable_type}\n"
            f"  File:            {filename}\n"
            f"  PII tokens redacted: {pii_count}\n"
            f"  Workflow ID:     {workflow_id}\n"
            f"  Clearance sent to: {gate.approver_email}\n"
            f"\nNext step: Once cleared, run:\n"
            f"  python claude_code_agent_ecosystem.py run_deliverable_review "
            f"--deliverable-id {deliverable_id}\n"
            f"{'='*70}\n"
        )

    def _review_scope_compliance(self, deliverable_text: str,
                                  deliverable_id: str, vendor_name: str,
                                  deliverable_type: str) -> Tuple[dict, int]:
        """PM Agent: check deliverable against AURA project scope."""
        template = REVIEW_TEMPLATES["pm_agent"]["scope_compliance_review"]
        prompt = template.format(
            deliverable_id=deliverable_id,
            vendor_name=vendor_name,
            deliverable_type=deliverable_type,
            review_date=datetime.now().strftime("%Y-%m-%d %H:%M"),
            deliverable_text=deliverable_text[:4000],  # truncate for token budget
        )
        output, call, confidence = self._call_claude(
            agent_name="pm_agent",
            system_prompt=(
                "You are the PM Agent for First Genesis. "
                "Your Phase 1 role is ONLY to review vendor deliverables — not to create them. "
                "Output valid JSON only. No explanations."
            ),
            user_message=prompt,
            scaffold_mode=True,
        )
        try:
            assessment = json.loads(output)
        except json.JSONDecodeError:
            assessment = {
                "scope_compliance_score": 1,
                "recommendation": "FLAG",
                "pm_notes": f"[PARSE ERROR — raw output: {output[:200]}]",
                "confidence": confidence,
            }
        return assessment, confidence

    def _review_requirements_compliance(self, deliverable_text: str,
                                         deliverable_id: str, vendor_name: str,
                                         deliverable_type: str) -> Tuple[dict, int]:
        """BA Agent: check deliverable against AURA requirements."""
        template = REVIEW_TEMPLATES["ba_agent"]["requirements_compliance_review"]
        prompt = template.format(
            deliverable_id=deliverable_id,
            vendor_name=vendor_name,
            deliverable_type=deliverable_type,
            review_date=datetime.now().strftime("%Y-%m-%d %H:%M"),
            deliverable_text=deliverable_text[:4000],
        )
        output, call, confidence = self._call_claude(
            agent_name="ba_agent",
            system_prompt=(
                "You are the BA Agent for First Genesis. "
                "Your Phase 1 role is ONLY to review vendor deliverables against requirements — "
                "not to create requirements. Output valid JSON only. No explanations."
            ),
            user_message=prompt,
            scaffold_mode=True,
        )
        try:
            assessment = json.loads(output)
        except json.JSONDecodeError:
            assessment = {
                "requirements_coverage_score": 1,
                "recommendation": "FLAG",
                "ba_notes": f"[PARSE ERROR — raw output: {output[:200]}]",
                "confidence": confidence,
            }
        return assessment, confidence

    def _review_quality_assessment(self, deliverable_text: str,
                                    deliverable_id: str, vendor_name: str,
                                    deliverable_type: str) -> Tuple[dict, int]:
        """QA Agent: assess quality and standards compliance."""
        template = REVIEW_TEMPLATES["qa_agent"]["quality_assessment_review"]
        prompt = template.format(
            deliverable_id=deliverable_id,
            vendor_name=vendor_name,
            deliverable_type=deliverable_type,
            review_date=datetime.now().strftime("%Y-%m-%d %H:%M"),
            deliverable_text=deliverable_text[:4000],
        )
        output, call, confidence = self._call_claude(
            agent_name="qa_agent",
            system_prompt=(
                "You are the QA Agent for First Genesis. "
                "Your Phase 1 role is ONLY to assess quality of vendor deliverables — "
                "not to create them. Output valid JSON only. No explanations."
            ),
            user_message=prompt,
            scaffold_mode=True,
        )
        try:
            assessment = json.loads(output)
        except json.JSONDecodeError:
            assessment = {
                "quality_score": 1,
                "recommendation": "FLAG",
                "qa_notes": f"[PARSE ERROR — raw output: {output[:200]}]",
                "confidence": confidence,
            }
        return assessment, confidence

    def run_vendor_deliverable_review(self, deliverable_id: str) -> str:
        """Phase 1 orchestrator — run all three agent reviews on a vendor deliverable.

        Sequence:
        1. Load deliverable from DB (must be in cleared/pending status)
        2. PM Agent: scope compliance review
        3. BA Agent: requirements compliance review
        4. QA Agent: quality assessment
        5. Build combined review report
        6. COMBINED_REVIEW gate → email full report to SME for final approval
        7. SME decision stored; used to improve future reviews via knowledge library
        """
        self.db.cursor.execute(
            "SELECT vendor_name, deliverable_type, filename, content_text, review_status "
            "FROM vendor_deliverables WHERE deliverable_id=?",
            (deliverable_id,)
        )
        row = self.db.cursor.fetchone()
        if not row:
            return f"ERROR: Deliverable '{deliverable_id}' not found."
        vendor_name, deliverable_type, filename, content_text, review_status = row

        if review_status == "pending_clearance":
            return (f"ERROR: Deliverable '{deliverable_id}' is still awaiting PII clearance. "
                    f"Approve the DOCUMENT_CLEARANCE gate first.")
        if review_status in ("sme_pending", "approved", "rejected"):
            return f"Deliverable '{deliverable_id}' already reviewed (status={review_status})."

        logger.info(f"Starting vendor review: {deliverable_id} ({vendor_name})")

        # Update status
        self.db.cursor.execute(
            "UPDATE vendor_deliverables SET review_status='under_review', updated_at=? "
            "WHERE deliverable_id=?",
            (datetime.now().isoformat(), deliverable_id)
        )
        self.db.conn.commit()

        deliverable_text = content_text or ""

        # ── Run all three agent reviews ──────────────────────────────────────
        logger.info("  PM Agent: scope compliance review...")
        pm_assessment, pm_conf = self._review_scope_compliance(
            deliverable_text, deliverable_id, vendor_name, deliverable_type)

        logger.info("  BA Agent: requirements compliance review...")
        ba_assessment, ba_conf = self._review_requirements_compliance(
            deliverable_text, deliverable_id, vendor_name, deliverable_type)

        logger.info("  QA Agent: quality assessment...")
        qa_assessment, qa_conf = self._review_quality_assessment(
            deliverable_text, deliverable_id, vendor_name, deliverable_type)

        # ── Persist individual assessments ───────────────────────────────────
        pm_score = pm_assessment.get("scope_compliance_score", 1)
        ba_score = ba_assessment.get("requirements_coverage_score", 1)
        qa_score = qa_assessment.get("quality_score", 1)
        overall_score = round((pm_score + ba_score + qa_score) / 3, 2)

        recommendations = [
            pm_assessment.get("recommendation", "FLAG"),
            ba_assessment.get("recommendation", "FLAG"),
            qa_assessment.get("recommendation", "FLAG"),
        ]
        if "REJECT" in recommendations:
            overall_rec = "REJECT"
        elif recommendations.count("APPROVE") == 3:
            overall_rec = "APPROVE"
        else:
            overall_rec = "FLAG"

        risks = [pm_assessment.get("timeline_risk", "MEDIUM"),
                 qa_assessment.get("risk_level", "MEDIUM")]
        overall_risk = "HIGH" if "HIGH" in risks else ("MEDIUM" if "MEDIUM" in risks else "LOW")

        self.db.cursor.execute("""
            UPDATE vendor_deliverables SET
                pm_assessment=?, pm_score=?,
                ba_assessment=?, ba_score=?,
                qa_assessment=?, qa_score=?,
                overall_score=?, overall_risk=?,
                recommendation=?, review_status='reviewed',
                updated_at=?
            WHERE deliverable_id=?
        """, (
            json.dumps(pm_assessment), pm_score,
            json.dumps(ba_assessment), ba_score,
            json.dumps(qa_assessment), qa_score,
            overall_score, overall_risk,
            overall_rec, datetime.now().isoformat(),
            deliverable_id
        ))
        self.db.conn.commit()

        # ── Build combined review report ─────────────────────────────────────
        flags = []
        if pm_assessment.get("out_of_scope_items"):
            flags.append(f"  OUT OF SCOPE: {', '.join(pm_assessment['out_of_scope_items'][:3])}")
        if pm_assessment.get("scope_gaps"):
            flags.append(f"  SCOPE GAPS: {', '.join(pm_assessment['scope_gaps'][:3])}")
        if ba_assessment.get("unsatisfied_requirements"):
            flags.append(f"  UNMET REQUIREMENTS: {', '.join(ba_assessment['unsatisfied_requirements'][:3])}")
        if qa_assessment.get("defects_found"):
            flags.append(f"  DEFECTS: {', '.join(qa_assessment['defects_found'][:3])}")
        if qa_assessment.get("scope_creep_detected"):
            flags.append(f"  SCOPE CREEP DETECTED: {qa_assessment.get('scope_creep_notes','')}")

        combined_report = (
            f"COMBINED AGENT REVIEW REPORT\n"
            f"{'='*60}\n"
            f"  Deliverable ID:  {deliverable_id}\n"
            f"  Vendor:          {vendor_name}\n"
            f"  Type:            {deliverable_type}\n"
            f"  File:            {filename}\n"
            f"  Review Date:     {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
            f"{'='*60}\n"
            f"AGENT SCORES:\n"
            f"  PM  Scope Compliance:        {pm_score}/5\n"
            f"  BA  Requirements Coverage:   {ba_score}/5\n"
            f"  QA  Quality Score:           {qa_score}/5\n"
            f"  Overall Score:               {overall_score}/5\n"
            f"  Overall Risk:                {overall_risk}\n"
            f"  Agent Recommendation:        {overall_rec}\n"
            f"{'='*60}\n"
            f"PM AGENT NOTES:\n  {pm_assessment.get('pm_notes','')}\n"
            f"  Timeline Risk: {pm_assessment.get('timeline_risk','')}\n"
            f"  Budget Impact: {pm_assessment.get('budget_impact','')}\n"
            f"\nBA AGENT NOTES:\n  {ba_assessment.get('ba_notes','')}\n"
            f"  Integration Concerns: {ba_assessment.get('integration_concerns',[])}\n"
            f"\nQA AGENT NOTES:\n  {qa_assessment.get('qa_notes','')}\n"
            f"  Completeness: {qa_assessment.get('completeness_pct',0)}%\n"
        )
        if flags:
            combined_report += f"\nFLAGS REQUIRING SME ATTENTION ({len(flags)}):\n"
            combined_report += "\n".join(flags) + "\n"

        combined_report += (
            f"\n{'='*60}\n"
            f"SME ACTION REQUIRED:\n"
            f"  Reply APPROVED to accept this deliverable.\n"
            f"  Reply REJECTED with feedback to return to vendor.\n"
            f"  All agent assessments stored and available for your review.\n"
            f"{'='*60}\n"
        )

        # Persist combined review
        self.db.cursor.execute(
            "UPDATE vendor_deliverables SET combined_review=?, updated_at=? "
            "WHERE deliverable_id=?",
            (combined_report, datetime.now().isoformat(), deliverable_id)
        )
        self.db.conn.commit()

        # ── COMBINED_REVIEW gate → SME approval ──────────────────────────────
        workflow_id = f"review_{deliverable_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        # Route low-confidence outputs to SME_REVIEW first
        min_conf = min(pm_conf, ba_conf, qa_conf)
        gate_name = StageGateName.SME_REVIEW if min_conf < 3 else StageGateName.COMBINED_REVIEW

        wf = self.stage_gate_manager.pause_at_gate(
            workflow_id=workflow_id,
            agent_name="pm_agent/ba_agent/qa_agent",
            project_name="AURA MVP",
            stage_gate_name=gate_name,
            content_pending_approval=combined_report
        )
        self.db.cursor.execute(
            "UPDATE vendor_deliverables SET workflow_id=?, review_status='sme_pending', "
            "updated_at=? WHERE deliverable_id=?",
            (workflow_id, datetime.now().isoformat(), deliverable_id)
        )
        self.db.conn.commit()

        # ── Knowledge library: log this review as a workflow pattern ─────────
        self.knowledge_library.cursor.execute("""
            INSERT INTO workflow_patterns
            (pattern_id, name, description, steps, success_rate, avg_duration_hours,
             use_count, last_used)
            VALUES (?, ?, ?, ?, ?, ?, 1, ?)
            ON CONFLICT(pattern_id) DO UPDATE SET use_count=use_count+1, last_used=?
        """, (
            f"vendor_review_{deliverable_type.lower().replace(' ','_')}",
            f"Vendor {deliverable_type} Review",
            f"3-agent review of vendor {deliverable_type} deliverable",
            f"intake → PII_clearance → PM_review → BA_review → QA_review → SME_approval",
            100.0, 1.0,
            datetime.now().isoformat(), datetime.now().isoformat()
        ))
        self.knowledge_library.conn.commit()

        gate = StageGateManager.STAGE_GATES[gate_name]
        return (
            f"\n{'='*70}\n"
            f"VENDOR DELIVERABLE REVIEW COMPLETE\n"
            f"{'='*70}\n"
            f"  Deliverable:  {filename} ({vendor_name})\n"
            f"  Overall Score: {overall_score}/5  |  Risk: {overall_risk}\n"
            f"  Recommendation: {overall_rec}\n"
            f"  Workflow ID:   {workflow_id}\n"
            f"  Review sent to SME: {gate.approver_email}\n"
            f"  CC: {', '.join(gate.cc_emails)}\n"
            f"\nAll agents have reviewed. Awaiting SME final decision.\n"
            f"{'='*70}\n"
        )

    def get_deliverable_queue(self) -> str:
        """Show all vendor deliverables and their review pipeline status."""
        self.db.cursor.execute("""
            SELECT deliverable_id, vendor_name, deliverable_type, filename,
                   review_status, overall_score, overall_risk, recommendation,
                   submitted_at, updated_at
            FROM vendor_deliverables ORDER BY submitted_at DESC
        """)
        rows = self.db.cursor.fetchall()
        lines = ["\n" + "="*70, "PHASE 1 — VENDOR DELIVERABLE QUEUE", "="*70]
        if not rows:
            lines.append("  No deliverables in queue.")
        for r in rows:
            did, vendor, dtype, fname, status, score, risk, rec, submitted, updated = r
            score_str = f"{score}/5" if score else "pending"
            risk_str = risk or "—"
            rec_str = rec or "—"
            lines.append(
                f"\n  [{did}]\n"
                f"    Vendor:      {vendor}\n"
                f"    Type:        {dtype}\n"
                f"    File:        {fname}\n"
                f"    Status:      {status}\n"
                f"    Score:       {score_str}  |  Risk: {risk_str}  |  Rec: {rec_str}\n"
                f"    Submitted:   {submitted}\n"
                f"    Last Update: {updated}"
            )

        # Alert on deliverables stuck > 48h without SME decision
        self.db.cursor.execute("""
            SELECT deliverable_id, vendor_name, submitted_at
            FROM vendor_deliverables WHERE review_status = 'sme_pending'
        """)
        alerts = []
        for did, vendor, submitted_at in self.db.cursor.fetchall():
            try:
                age_h = (datetime.now() - datetime.fromisoformat(submitted_at)).total_seconds() / 3600
                if age_h > 48:
                    alerts.append(f"  [STALE] {did} ({vendor}) waiting {age_h:.1f}h for SME decision")
            except (ValueError, TypeError):
                pass
        if alerts:
            lines.append(f"\nSTALE ALERTS ({len(alerts)}):")
            lines.extend(alerts)
        lines.append("="*70 + "\n")
        return "\n".join(lines)

    def submit_sme_correction(self, agent_name: str, original_output_hash: str,
                              original_snippet: str, corrected_content: str,
                              correction_category: str, corrector_name: str,
                              weight: int = 1) -> str:
        """Record an SME correction for an agent output.
        The correction will be injected as a few-shot example in that agent's
        future system prompts, shaping its behaviour without retraining.
        """
        correction_id = self.knowledge_library.save_sme_correction(
            agent_name=agent_name,
            original_output_hash=original_output_hash,
            original_snippet=original_snippet,
            corrected_content=corrected_content,
            correction_category=correction_category,
            corrector_name=corrector_name,
            weight=weight
        )
        return (f"\nSME correction recorded:\n"
                f"  Correction ID: {correction_id}\n"
                f"  Agent:         {agent_name}\n"
                f"  Category:      {correction_category}\n"
                f"  Corrector:     {corrector_name}\n"
                f"  Weight:        {weight}\n"
                f"This correction will auto-inject into future {agent_name} calls.\n")




# ============================================================================
# DASHBOARD: KNOWLEDGE LIBRARY
# ============================================================================
class KnowledgeLibrary:
    """SQLite database for agent learning, workflow patterns, and lessons."""

    def __init__(self, db_path: str = None):
        if db_path is None:
            db_path = os.environ.get("FG_KNOWLEDGE_DB_PATH", "/home/claude/fg_knowledge.db")
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path, check_same_thread=False)
        self._init_tables()

    def _init_tables(self):
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS skills (
                skill_id TEXT PRIMARY KEY,
                agent_name TEXT,
                skill_name TEXT,
                description TEXT,
                success_count INTEGER DEFAULT 0,
                error_count INTEGER DEFAULT 0,
                avg_execution_time REAL,
                skill_level TEXT,
                last_used TEXT,
                template TEXT,
                created_at TEXT
            );
            CREATE TABLE IF NOT EXISTS workflow_patterns (
                workflow_id TEXT PRIMARY KEY,
                workflow_name TEXT,
                agent_sequence TEXT,
                success_count INTEGER DEFAULT 0,
                avg_duration_seconds REAL,
                avg_cost_usd REAL,
                template TEXT,
                created_at TEXT,
                last_used TEXT
            );
            CREATE TABLE IF NOT EXISTS lessons_learned (
                lesson_id TEXT PRIMARY KEY,
                workflow_id TEXT,
                lesson_title TEXT,
                lesson_content TEXT,
                category TEXT,
                applicable_agents TEXT,
                created_at TEXT,
                usage_count INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS agent_comm_log (
                message_id TEXT PRIMARY KEY,
                from_agent TEXT,
                to_agent TEXT,
                message_type TEXT,
                content TEXT,
                timestamp TEXT,
                status TEXT
            );

            CREATE TABLE IF NOT EXISTS sme_corrections (
                correction_id TEXT PRIMARY KEY,
                agent_name TEXT NOT NULL,
                original_output_hash TEXT NOT NULL,
                original_snippet TEXT NOT NULL,
                corrected_content TEXT NOT NULL,
                correction_category TEXT NOT NULL,
                corrector_name TEXT NOT NULL,
                weight INTEGER DEFAULT 1,
                created_at TEXT NOT NULL,
                usage_count INTEGER DEFAULT 0
            );
        """)
        self.conn.commit()

    def save_skill(self, skill: AgentSkill):
        self.conn.execute("""
            INSERT OR REPLACE INTO skills
            (skill_id, agent_name, skill_name, description, success_count,
             error_count, avg_execution_time, skill_level, last_used, template, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            skill.skill_id, skill.agent_name, skill.skill_name, skill.description,
            skill.success_count, skill.error_count, skill.avg_execution_time,
            skill.skill_level.name, skill.last_used, json.dumps(skill.template),
            datetime.now().isoformat()
        ))
        self.conn.commit()

    def get_agent_skills(self, agent_name: str) -> List[AgentSkill]:
        cursor = self.conn.execute(
            "SELECT * FROM skills WHERE agent_name = ?", (agent_name,)
        )
        skills = []
        for row in cursor.fetchall():
            skills.append(AgentSkill(
                skill_id=row[0], agent_name=row[1], skill_name=row[2],
                description=row[3], success_count=row[4], error_count=row[5],
                avg_execution_time=row[6], skill_level=SkillLevel[row[7]],
                last_used=row[8], template=json.loads(row[9]) if row[9] else {}
            ))
        return skills

    def save_workflow_pattern(self, workflow: WorkflowExecution):
        workflow_name = "-".join(workflow.agent_sequence)
        self.conn.execute("""
            INSERT OR REPLACE INTO workflow_patterns
            (workflow_id, workflow_name, agent_sequence, success_count,
             avg_duration_seconds, avg_cost_usd, template, created_at, last_used)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            workflow.workflow_id, workflow_name,
            json.dumps(workflow.agent_sequence), 1,
            workflow.duration_seconds, workflow.cost_usd,
            json.dumps({"input": workflow.input_data, "agent_sequence": workflow.agent_sequence}),
            datetime.now().isoformat(), datetime.now().isoformat()
        ))
        self.conn.commit()

    def get_workflow_patterns(self) -> List[Dict]:
        cursor = self.conn.execute(
            "SELECT * FROM workflow_patterns ORDER BY success_count DESC"
        )
        patterns = []
        for row in cursor.fetchall():
            patterns.append({
                "workflow_id": row[0], "workflow_name": row[1],
                "agent_sequence": json.loads(row[2]), "success_count": row[3],
                "avg_duration_seconds": row[4], "avg_cost_usd": row[5],
                "template": json.loads(row[6]) if row[6] else None
            })
        return patterns

    def save_lesson_learned(self, workflow_id: str, lesson_title: str,
                            lesson_content: str, category: str,
                            applicable_agents: List[str]):
        self.conn.execute("""
            INSERT INTO lessons_learned
            (lesson_id, workflow_id, lesson_title, lesson_content, category,
             applicable_agents, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            str(uuid.uuid4()), workflow_id, lesson_title, lesson_content,
            category, json.dumps(applicable_agents), datetime.now().isoformat()
        ))
        self.conn.commit()

    # ── SME Corrections ───────────────────────────────────────────────────────

    def save_sme_correction(self, agent_name: str, original_output_hash: str,
                            original_snippet: str, corrected_content: str,
                            correction_category: str, corrector_name: str,
                            weight: int = 1):
        """Store an SME correction to inject as a few-shot example in future calls."""
        correction_id = str(uuid.uuid4())
        self.conn.execute("""
            INSERT INTO sme_corrections
            (correction_id, agent_name, original_output_hash, original_snippet,
             corrected_content, correction_category, corrector_name, weight, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            correction_id, agent_name, original_output_hash,
            original_snippet[:500], corrected_content[:1000],
            correction_category, corrector_name, weight,
            datetime.now().isoformat()
        ))
        self.conn.commit()
        logger.info(f"SME correction saved for {agent_name} by {corrector_name}")
        return correction_id

    def get_sme_corrections(self, agent_name: str, limit: int = 3) -> List[Dict]:
        """Retrieve top-weighted corrections for an agent to inject into prompts."""
        cursor = self.conn.execute("""
            SELECT correction_id, agent_name, original_snippet, corrected_content,
                   correction_category, corrector_name, weight, usage_count, created_at
            FROM sme_corrections
            WHERE agent_name = ?
            ORDER BY weight DESC, usage_count ASC
            LIMIT ?
        """, (agent_name, limit))
        corrections = []
        for row in cursor.fetchall():
            corrections.append({
                "correction_id": row[0], "agent_name": row[1],
                "original_snippet": row[2], "corrected_content": row[3],
                "category": row[4], "corrector_name": row[5],
                "weight": row[6], "usage_count": row[7], "created_at": row[8]
            })
        if corrections:
            ids = [c["correction_id"] for c in corrections]
            self.conn.execute(
                f"UPDATE sme_corrections SET usage_count = usage_count + 1 "
                f"WHERE correction_id IN ({','.join('?'*len(ids))})", ids
            )
            self.conn.commit()
        return corrections

    def get_lessons_learned(self, category: Optional[str] = None) -> List[Dict]:  # type: ignore[override]
        if category:
            cursor = self.conn.execute(
                "SELECT * FROM lessons_learned WHERE category = ? ORDER BY created_at DESC",
                (category,)
            )
        else:
            cursor = self.conn.execute(
                "SELECT * FROM lessons_learned ORDER BY created_at DESC"
            )
        lessons = []
        for row in cursor.fetchall():
            lessons.append({
                "lesson_id": row[0], "workflow_id": row[1], "lesson_title": row[2],
                "lesson_content": row[3], "category": row[4],
                "applicable_agents": json.loads(row[5]), "created_at": row[6],
                "usage_count": row[7]
            })
        return lessons


# ============================================================================
# DASHBOARD: AGENT COMMUNICATION BUS
# ============================================================================
class AgentCommunicationBus:
    """Message bus for inter-agent communication."""

    MAX_HISTORY = 10_000

    def __init__(self, knowledge_library: KnowledgeLibrary):
        self.knowledge_library = knowledge_library
        self.message_queue: queue.Queue = queue.Queue()
        self.message_history: List[AgentMessage] = []
        self.agents: Dict[str, AgentRecord] = {}

    def register_agent(self, agent: AgentRecord):
        self.agents[agent.name] = agent

    def send_message(self, from_agent: str, to_agent: str,
                     message_type: MessageType, content: Dict) -> str:
        message_id = str(uuid.uuid4())
        msg = AgentMessage(
            id=message_id, from_agent=from_agent, to_agent=to_agent,
            message_type=message_type, content=content,
            timestamp=datetime.now().isoformat(), status="sent"
        )
        self.message_queue.put(msg)
        self.message_history.append(msg)
        if len(self.message_history) > self.MAX_HISTORY:
            self.message_history = self.message_history[-self.MAX_HISTORY:]
        return message_id

    def get_messages_for_agent(self, agent_name: str) -> List[AgentMessage]:
        return [m for m in self.message_history if m.to_agent == agent_name]

    def get_conversation(self, agent1: str, agent2: str) -> List[AgentMessage]:
        return [
            m for m in self.message_history
            if (m.from_agent == agent1 and m.to_agent == agent2) or
               (m.from_agent == agent2 and m.to_agent == agent1)
        ]


# ============================================================================
# DASHBOARD: SKILL COMPOUNDING ENGINE
# ============================================================================
class SkillCompoundingEngine:
    """Tracks and improves agent skills over time."""

    def __init__(self, knowledge_library: KnowledgeLibrary):
        self.knowledge_library = knowledge_library

    def record_success(self, agent_name: str, skill_name: str,
                       execution_time: float, output_quality: float = 1.0):
        skills = self.knowledge_library.get_agent_skills(agent_name)
        skill = next((s for s in skills if s.skill_name == skill_name), None)
        if not skill:
            skill = AgentSkill(
                skill_id=str(uuid.uuid4()), agent_name=agent_name,
                skill_name=skill_name, description=f"Skill learned by {agent_name}",
                success_count=0, error_count=0, avg_execution_time=0,
                skill_level=SkillLevel.NOVICE, last_used=datetime.now().isoformat(),
                template={}
            )
        skill.success_count += 1
        skill.last_used = datetime.now().isoformat()
        if skill.success_count >= 50:
            skill.skill_level = SkillLevel.EXPERT
        elif skill.success_count >= 20:
            skill.skill_level = SkillLevel.ADVANCED
        elif skill.success_count >= 5:
            skill.skill_level = SkillLevel.INTERMEDIATE
        if skill.avg_execution_time == 0:
            skill.avg_execution_time = execution_time
        else:
            skill.avg_execution_time = (
                (skill.avg_execution_time * (skill.success_count - 1) + execution_time)
                / skill.success_count
            )
        self.knowledge_library.save_skill(skill)

    def record_error(self, agent_name: str, skill_name: str, error: str):
        skills = self.knowledge_library.get_agent_skills(agent_name)
        skill = next((s for s in skills if s.skill_name == skill_name), None)
        if skill:
            skill.error_count += 1
            self.knowledge_library.save_skill(skill)

    def get_agent_improvement(self, agent_name: str) -> Dict:
        skills = self.knowledge_library.get_agent_skills(agent_name)
        total_successes = sum(s.success_count for s in skills)
        total_errors = sum(s.error_count for s in skills)
        avg_level = sum(s.skill_level.value for s in skills) / len(skills) if skills else 0
        total = total_successes + total_errors
        return {
            "agent_name": agent_name,
            "total_skills": len(skills),
            "total_successes": total_successes,
            "total_errors": total_errors,
            "success_rate": total_successes / total if total > 0 else 0,
            "avg_skill_level": avg_level,
            "skills": [asdict(s) for s in skills]
        }


# ============================================================================
# DASHBOARD: STATE MANAGER & COMMAND CENTER
# ============================================================================
class DashboardStateManager:
    """Manages complete state for real-time dashboard display."""

    def __init__(self):
        self.knowledge_library = KnowledgeLibrary()
        self.communication_bus = AgentCommunicationBus(self.knowledge_library)
        self.skill_engine = SkillCompoundingEngine(self.knowledge_library)
        self.agents: Dict[str, AgentRecord] = {}
        self.last_update = datetime.now().isoformat()

    def register_agent(self, agent: AgentRecord):
        self.agents[agent.name] = agent
        self.communication_bus.register_agent(agent)

    def update_agent_status(self, agent_name: str, status: AgentStatus,
                            current_task: Optional[str] = None):
        if agent_name in self.agents:
            self.agents[agent_name].status = status
            self.agents[agent_name].current_task = current_task
            self.agents[agent_name].last_activity = datetime.now().isoformat()
            self.last_update = datetime.now().isoformat()

    def get_dashboard_data(self) -> Dict:
        agents_data = []
        for agent_name, agent in self.agents.items():
            agents_data.append({
                "name": agent.name, "role": agent.role,
                "status": agent.status.value, "current_task": agent.current_task,
                "skill_level": agent.skill_level.value,
                "success_count": agent.success_count, "error_count": agent.error_count,
                "last_activity": agent.last_activity,
                "active_workflows": len(agent.active_workflows)
            })
        recent_messages = [
            {"from": m.from_agent, "to": m.to_agent,
             "type": m.message_type.value, "timestamp": m.timestamp}
            for m in self.communication_bus.message_history[-100:]
        ]
        patterns = self.knowledge_library.get_workflow_patterns()
        return {
            "timestamp": self.last_update,
            "agents": agents_data,
            "recent_communications": recent_messages,
            "workflow_patterns": patterns,
            "lessons_learned_count": len(self.knowledge_library.get_lessons_learned()),
            "total_skills": sum(
                len(self.skill_engine.get_agent_improvement(name)["skills"])
                for name in self.agents
            )
        }


class CommandCenter:
    """API surface for command center and dashboard queries."""

    def __init__(self, state_manager: DashboardStateManager):
        self.state = state_manager

    def get_dashboard(self) -> Dict:
        return self.state.get_dashboard_data()

    def get_agent_details(self, agent_name: str) -> Dict:
        agent = self.state.agents.get(agent_name)
        if not agent:
            return {"error": f"Agent {agent_name} not found"}
        improvement = self.state.skill_engine.get_agent_improvement(agent_name)
        recent_messages = self.state.communication_bus.get_messages_for_agent(agent_name)
        return {
            "agent": asdict(agent),
            "improvement_metrics": improvement,
            "recent_messages": [
                {"from": m.from_agent, "type": m.message_type.value,
                 "timestamp": m.timestamp}
                for m in recent_messages[-20:]
            ]
        }

    def get_workflow_patterns(self) -> List[Dict]:
        return self.state.knowledge_library.get_workflow_patterns()

    def get_lessons_learned(self, category: Optional[str] = None) -> List[Dict]:
        return self.state.knowledge_library.get_lessons_learned(category)

    def get_conversation_log(self, agent1: str, agent2: str) -> List[Dict]:
        messages = self.state.communication_bus.get_conversation(agent1, agent2)
        return [
            {"from": m.from_agent, "to": m.to_agent, "type": m.message_type.value,
             "content": m.content, "timestamp": m.timestamp}
            for m in messages
        ]

    def get_skill_progression(self, agent_name: str) -> Dict:
        return self.state.skill_engine.get_agent_improvement(agent_name)

    def send_agent_message(self, from_agent: str, to_agent: str,
                           message_type: str, content: Dict) -> str:
        msg_type = MessageType[message_type.upper()]
        return self.state.communication_bus.send_message(
            from_agent, to_agent, msg_type, content
        )


# ============================================================================
# DASHBOARD: DEMO / SMOKE TEST
# ============================================================================
def demo_dashboard():
    """Demonstrate dashboard system with simulated agent activity."""

    print("\n" + "=" * 70)
    print("AGENT DASHBOARD & COMMAND CENTER DEMO")
    print("=" * 70 + "\n")

    state = DashboardStateManager()
    command_center = CommandCenter(state)

    # Register agents
    agents = [
        AgentRecord("pm_1",     "PM Agent",     "Project Manager",   AgentStatus.IDLE,     None,           SkillLevel.INTERMEDIATE, 5, 0, datetime.now().isoformat(), []),
        AgentRecord("ba_1",     "BA Agent",     "Business Analyst",  AgentStatus.THINKING, "Requirements", SkillLevel.INTERMEDIATE, 3, 0, datetime.now().isoformat(), []),
        AgentRecord("qa_1",     "QA Agent",     "Quality Assurance", AgentStatus.IDLE,     None,           SkillLevel.NOVICE,       2, 1, datetime.now().isoformat(), []),
        AgentRecord("vendor_1", "Vendor Agent", "Vendor Manager",    AgentStatus.IDLE,     None,           SkillLevel.NOVICE,       1, 0, datetime.now().isoformat(), []),
        AgentRecord("mgr_1",    "Manager Agent","Portfolio Manager",  AgentStatus.IDLE,     None,           SkillLevel.INTERMEDIATE, 4, 0, datetime.now().isoformat(), []),
    ]
    for a in agents:
        state.register_agent(a)

    # Simulate inter-agent communication
    print("1. SIMULATING AGENT COMMUNICATION:")
    print("-" * 70)
    msg_id = state.communication_bus.send_message(
        "PM Agent", "BA Agent", MessageType.DELEGATE,
        {"task": "Extract requirements from AURA design session"}
    )
    print(f"   PM Agent → BA Agent: 'Extract requirements'  [ID: {msg_id[:8]}...]")
    state.communication_bus.send_message(
        "BA Agent", "PM Agent", MessageType.PROVIDE_OUTPUT,
        {"requirements": ["FR1: Silhouette processing", "FR2: 3D mesh export", "FR3: Actor model tagging"]}
    )
    print(f"   BA Agent → PM Agent: Provided 3 requirements")
    state.communication_bus.send_message(
        "PM Agent", "QA Agent", MessageType.REQUEST_INPUT,
        {"task": "Review charter quality before gate submission"}
    )
    print(f"   PM Agent → QA Agent: 'Review charter quality'\n")

    # Record skill executions
    print("2. RECORDING SKILL IMPROVEMENTS:")
    print("-" * 70)
    state.skill_engine.record_success("PM Agent", "create_charter",          120.5, 0.95)
    state.skill_engine.record_success("PM Agent", "create_charter",          115.0, 0.98)
    state.skill_engine.record_success("BA Agent", "extract_requirements",    180.0, 0.92)
    state.skill_engine.record_success("QA Agent", "audit_deliverable",       90.0,  0.88)
    state.skill_engine.record_success("Manager Agent", "portfolio_review",   60.0,  0.96)
    print("   ✅ PM Agent:      create_charter (×2, avg 117.75s)")
    print("   ✅ BA Agent:      extract_requirements (×1, avg 180s)")
    print("   ✅ QA Agent:      audit_deliverable (×1, avg 90s)")
    print("   ✅ Manager Agent: portfolio_review (×1, avg 60s)\n")

    # Show progression
    print("3. AGENT SKILL PROGRESSION:")
    print("-" * 70)
    for agent_name in ["PM Agent", "BA Agent", "QA Agent"]:
        m = state.skill_engine.get_agent_improvement(agent_name)
        print(f"   {agent_name}: {m['total_skills']} skill(s) | "
              f"Success rate {m['success_rate']*100:.0f}% | "
              f"Avg level {m['avg_skill_level']:.1f}/4")

    # Save workflow pattern
    print("\n4. CAPTURING WORKFLOW PATTERNS:")
    print("-" * 70)
    wf = WorkflowExecution(
        workflow_id=str(uuid.uuid4()),
        agent_sequence=["PM Agent", "BA Agent", "QA Agent"],
        start_time=datetime.now().isoformat(),
        end_time=datetime.now().isoformat(),
        duration_seconds=300, success=True,
        input_data={"project": "AURA MVP"},
        output_data={"charter_approved": True},
        errors=[], approvals_needed=1, approvals_completed=1, cost_usd=0.03
    )
    state.knowledge_library.save_workflow_pattern(wf)
    print("   ✅ Pattern saved: PM Agent → BA Agent → QA Agent")

    # Save lesson learned
    print("\n5. CAPTURING LESSONS LEARNED:")
    print("-" * 70)
    state.knowledge_library.save_lesson_learned(
        wf.workflow_id,
        "Charter quality improves with design session review",
        "When BA Agent reviews design sessions first, PM Agent creates better charters",
        "process_optimization", ["PM Agent", "BA Agent"]
    )
    print("   ✅ Lesson: Charter quality improves with design session review")

    # Dashboard snapshot
    print("\n6. DASHBOARD SNAPSHOT:")
    print("-" * 70)
    dashboard = command_center.get_dashboard()
    print(f"   Agents Online:           {len(dashboard['agents'])}")
    for a in dashboard['agents']:
        print(f"     • {a['name']:<16} {a['status']:<20} (skill lvl {a['skill_level']}/4)")
    print(f"   Recent Communications:   {len(dashboard['recent_communications'])}")
    print(f"   Workflow Patterns:        {len(dashboard['workflow_patterns'])}")
    print(f"   Lessons Learned:          {dashboard['lessons_learned_count']}")
    print(f"   Total Skills Tracked:     {dashboard['total_skills']}")

    print("\n" + "=" * 70)
    print("DASHBOARD & COMMAND CENTER CAPABILITIES")
    print("=" * 70)
    print("""
  ✅ AGENT MONITORING    — Real-time status, task, success/error counts
  ✅ INTER-AGENT COMMS   — Message bus (6 types: initiate/delegate/feedback…)
  ✅ SKILL COMPOUNDING   — Novice → Intermediate → Advanced → Expert (50+)
  ✅ KNOWLEDGE LIBRARY   — Reusable workflow patterns + lessons learned
  ✅ PERFORMANCE METRICS — Success rate, avg skill level, progression
  ✅ COMMAND CENTER API  — Dashboard, agent details, conversation log
    """)


# ============================================================================
# TOKEN STRATEGY: PRICING & COST MODELS
# ============================================================================
class TokenPricing:
    """claude-opus-4-6 token pricing."""
    input_cost_per_1m = 5.00
    output_cost_per_1m = 25.00
    cache_write_cost_per_1m = 0.50
    cache_read_cost_per_1m = 0.50

    @staticmethod
    def calculate_cost(input_tokens: int, output_tokens: int,
                       cache_created: int = 0, cache_read: int = 0) -> float:
        return (
            (input_tokens   / 1_000_000) * TokenPricing.input_cost_per_1m +
            (output_tokens  / 1_000_000) * TokenPricing.output_cost_per_1m +
            (cache_created  / 1_000_000) * TokenPricing.cache_write_cost_per_1m +
            (cache_read     / 1_000_000) * TokenPricing.cache_read_cost_per_1m
        )


class AgentCostModel:
    """Realistic token costs for each agent type."""

    PM_AGENT     = {"name": "PM Agent",           "input_tokens": 3500, "output_tokens": 1200, "calls_per_day": 2, "description": "Project setup, WBS, status tracking"}
    BA_AGENT     = {"name": "BA Agent",           "input_tokens": 7000, "output_tokens": 1800, "calls_per_day": 3, "description": "Requirements, design sessions, data intake, schema extraction, dependency mapping"}
    QA_AGENT     = {"name": "QA Agent",           "input_tokens": 8000, "output_tokens": 2000, "calls_per_day": 1, "description": "Pre-delivery audit, scope creep detection"}
    VENDOR_AGENT = {"name": "Vendor Agent",       "input_tokens": 4000, "output_tokens":  800, "calls_per_day": 1, "description": "Partner SLA tracking, performance monitoring"}
    MANAGER_AGENT= {"name": "Manager Agent",      "input_tokens": 5000, "output_tokens": 1500, "calls_per_day": 1, "description": "Portfolio dashboard, orchestration"}
    ALL_AGENTS = [PM_AGENT, BA_AGENT, QA_AGENT, VENDOR_AGENT, MANAGER_AGENT]

    @staticmethod
    def cost_per_agent(agent: dict) -> float:
        return TokenPricing.calculate_cost(agent["input_tokens"], agent["output_tokens"])

    @staticmethod
    def daily_cost_per_agent(agent: dict) -> float:
        return AgentCostModel.cost_per_agent(agent) * agent["calls_per_day"]

    @staticmethod
    def total_daily_cost() -> float:
        return sum(AgentCostModel.daily_cost_per_agent(a) for a in AgentCostModel.ALL_AGENTS)


class TokenBudgetModel:
    """Executive budget allocation model (reporting-focused)."""

    DAILY_BUDGET = 5.00
    CONTINGENCY_PCT = 0.20

    @staticmethod
    def daily_agent_cost() -> float:
        return AgentCostModel.total_daily_cost()

    @staticmethod
    def budget_allocation() -> dict:
        agent_cost = TokenBudgetModel.daily_agent_cost()
        contingency = TokenBudgetModel.DAILY_BUDGET * TokenBudgetModel.CONTINGENCY_PCT
        return {
            "Daily Budget":          f"${TokenBudgetModel.DAILY_BUDGET:.2f}",
            "Agent Execution Cost":  f"${agent_cost:.2f}",
            "Contingency (20%)":     f"${contingency:.2f}",
            "Safety Buffer":         f"${TokenBudgetModel.DAILY_BUDGET - agent_cost - contingency:.2f}",
            "Headroom Multiplier":   f"{TokenBudgetModel.DAILY_BUDGET / agent_cost:.1f}x",
        }

    @staticmethod
    def monthly_cost(days: int = 30) -> float:
        return TokenBudgetModel.daily_agent_cost() * days


class OptimizationTechniques:
    """3 token optimization techniques with impact analysis."""

    @staticmethod
    def technique_1_prompt_precision() -> dict:
        verbose_cost   = TokenPricing.calculate_cost(1500, 2400)
        precision_cost = TokenPricing.calculate_cost(1500,  800)
        savings     = verbose_cost - precision_cost
        savings_pct = (savings / verbose_cost) * 100
        return {
            "technique":        "Prompt Precision",
            "description":      "Use exact output format (JSON, tables). No explanations.",
            "example":          "Verbose: 2,400 output tokens  ->  Precision JSON: 800 output tokens",
            "savings_per_call": f"${savings:.4f} ({savings_pct:.0f}%)",
            "monthly_savings":  f"${savings * 60:.2f} (60 charter calls/month)",
            "impact":           "MAJOR (67% output token reduction)",
        }

    @staticmethod
    def technique_2_caching() -> dict:
        lib_size = 30_000
        no_cache_monthly = TokenPricing.calculate_cost(4500 * 6, 0) * 30
        cache_calls      = TokenPricing.calculate_cost( 500 * 6, 0) * 30
        cache_write      = TokenPricing.calculate_cost(lib_size, 0, cache_created=lib_size)
        monthly_savings  = no_cache_monthly - (cache_calls + cache_write)
        return {
            "technique":       "Caching Templates",
            "description":     "Load template library once, reuse across all agent calls.",
            "example":         "60+ templates (30K tokens) cached -> reused every call",
            "cache_cost":      f"${cache_write:.2f} (one-time)",
            "monthly_savings": f"${monthly_savings:.2f}",
            "breakeven":       "2 days",
            "impact":          "MAJOR (55% input token reduction)",
        }

    @staticmethod
    def technique_3_scheduled_batching() -> dict:
        redundant = TokenPricing.calculate_cost(3500, 0)  # 1 avoided call/day
        return {
            "technique":             "Scheduled Batching",
            "description":           "Run agents on fixed schedule (6 AM, 9 AM, 5 PM). Batch output for review.",
            "example":               "3 on-demand calls (same data)  ->  2 scheduled calls",
            "redundant_tokens_daily":"3,500 tokens",
            "daily_savings":         f"${redundant:.4f}",
            "monthly_savings":       f"${redundant * 30:.2f}",
            "impact":                "MODERATE (33% redundant call elimination)",
        }

    @staticmethod
    def total_optimization_impact() -> dict:
        baseline  = TokenBudgetModel.daily_agent_cost()
        optimized = baseline * 0.25   # 75% combined reduction (conservative)
        return {
            "baseline_daily_cost":  f"${baseline:.4f}",
            "optimized_daily_cost": f"${optimized:.4f}",
            "daily_savings":        f"${baseline - optimized:.4f}",
            "monthly_savings":      f"${(baseline - optimized) * 30:.2f}",
            "combined_impact":      "75% overall token reduction",
            "headroom":             f"Can support {5 / optimized:.0f}x more agents before budget stress",
        }


# ============================================================================
# TOKEN STRATEGY: DISPLAY FUNCTIONS
# ============================================================================
def _print_header(title: str, width: int = 70):
    print(f"\n{'='*width}\n{title.center(width)}\n{'='*width}\n")


def show_budget_model():
    _print_header("BUDGET MODEL: Daily Allocation")
    for key, value in TokenBudgetModel.budget_allocation().items():
        print(f"{key:.<40} {value:>20}")
    print(f"\n{'Status:':<40} Within budget ({TokenBudgetModel.DAILY_BUDGET / TokenBudgetModel.daily_agent_cost():.0f}x headroom)")


def show_cost_breakdown():
    _print_header("AGENT COST BREAKDOWN")
    print(f"{'Agent':<25} {'Per Call':>15} {'Daily Cost':>15}")
    print("-" * 55)
    total = 0.0
    for agent in AgentCostModel.ALL_AGENTS:
        per_call = AgentCostModel.cost_per_agent(agent)
        daily    = AgentCostModel.daily_cost_per_agent(agent)
        total   += daily
        print(f"{agent['name']:<25} ${per_call:>14.4f} ${daily:>14.4f}")
    print("-" * 55)
    print(f"{'TOTAL DAILY':<25} {'':>15} ${total:>14.4f}")
    print(f"\nMonthly Cost (30 days): ${TokenBudgetModel.monthly_cost():.2f}")
    print("Equivalent FTE Avoided: $15,000+ (monthly salary)")


def show_optimization_impact():
    _print_header("OPTIMIZATION TECHNIQUES: Token Savings")
    for i, technique_fn in enumerate([
        OptimizationTechniques.technique_1_prompt_precision,
        OptimizationTechniques.technique_2_caching,
        OptimizationTechniques.technique_3_scheduled_batching,
    ], 1):
        t = technique_fn()
        print(f"TECHNIQUE {i}: {t['technique']}")
        for key, val in t.items():
            if key != "technique":
                print(f"  {key.replace('_', ' ').title()}: {val}")
        print()

    combined = OptimizationTechniques.total_optimization_impact()
    print("COMBINED IMPACT (All 3 Techniques):")
    for key, val in combined.items():
        print(f"  {key.replace('_', ' ').title()}: {val}")


def show_cost_projections():
    _print_header("MONTHLY COST PROJECTIONS")
    scenarios = {
        "Aura Only (1 project)":       0.5,
        "Full Portfolio (5 projects)": 1.0,
        "Scaled (10+ projects)":       1.5,
    }
    for name, factor in scenarios.items():
        cost = TokenBudgetModel.monthly_cost() * factor
        flag = "OK" if cost <= 150 else "REVIEW" if cost <= 300 else "UPGRADE"
        print(f"  [{flag}] {name}: ${cost:.2f}")


def show_executive_summary():
    _print_header("EXECUTIVE SUMMARY: Token Strategy for First Genesis")
    agent_cost = TokenBudgetModel.daily_agent_cost()
    metrics = {
        "Daily Budget":                     "$5.00 USD",
        "Agent Execution Cost":             f"${agent_cost:.4f} USD",
        "Budget Headroom":                  f"{TokenBudgetModel.DAILY_BUDGET / agent_cost:.0f}x",
        "Monthly Cost (Full Portfolio)":    f"${TokenBudgetModel.monthly_cost():.2f} USD",
        "Token Optimization":               "75% reduction via 3 techniques",
        "Autonomy Level":                   "80% (agents run unsupervised)",
        "Human Approval Time (per workflow)": "2-5 minutes",
        "Setup Time":                       "3 weeks (foundation -> integration -> go-live)",
    }
    for key, val in metrics.items():
        print(f"  {key}: {val}")

    print("""
OPTIMIZATION TECHNIQUES:
  1. Prompt Precision:      67% output token reduction
  2. Caching Templates:     55% input token reduction
  3. Scheduled Batching:    33% redundant call elimination
  Combined:                 75% overall token reduction

RISK MITIGATION:
  Hard budget stops (no overage possible)
  Frozen facts validation (prevents hallucinations)
  Human-in-the-loop gates (approval before delivery)
  Workflow persistence (audit trail)
  Timeout escalation (auto-escalate if no approval)

IMPLEMENTATION TIMELINE:
  Week 1: Foundation (guardrail code, templates, API)
  Week 2: Integration (agents built & tested)
  Week 3: Scheduler (automation live)

STATUS: READY FOR DEPLOYMENT
""")


def show_token_dashboard():
    _print_header("TOKEN STRATEGY DASHBOARD", width=80)

    print("BUDGET STATUS:")
    for key, val in TokenBudgetModel.budget_allocation().items():
        print(f"  {key}: {val}")

    print("\nAGENT COSTS (Daily):")
    total = 0.0
    for agent in AgentCostModel.ALL_AGENTS:
        daily = AgentCostModel.daily_cost_per_agent(agent)
        total += daily
        print(f"  {agent['name']}: ${daily:.4f} ({agent['calls_per_day']} calls)")
    print(f"  {'─'*40}")
    print(f"  TOTAL: ${total:.4f}")

    print("\nOPTIMIZATION POTENTIAL:")
    combined = OptimizationTechniques.total_optimization_impact()
    for key, val in combined.items():
        print(f"  {key.replace('_', ' ').title()}: {val}")

    print("\nMONTHLY PROJECTIONS:")
    for name, factor in [("Aura Only", 0.5), ("Full Portfolio", 1.0), ("Scaled 10+", 1.5)]:
        cost = TokenBudgetModel.monthly_cost() * factor
        flag = "OK    " if cost <= 150 else "REVIEW" if cost <= 300 else "UPGRADE"
        print(f"  [{flag}] {name}: ${cost:.2f}")

    print("\nRECOMMENDATIONS:")
    recs = [
        "Use all 3 optimization techniques (75% savings)",
        "Implement email approval gates (human-in-the-loop)",
        "Schedule agents on fixed timeline (6 AM, 9 AM, 5 PM)",
        "Cache template library at startup (one-time cost)",
        "Monitor headroom monthly (decision point at >$300/month)",
    ]
    for i, rec in enumerate(recs, 1):
        print(f"  {i}. {rec}")
    print("\n" + "=" * 80)


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================
def main():
    if len(sys.argv) < 2:
        print("Usage: python claude_code_agent_ecosystem.py [command]")
        print()
        print("╔══════════════════════════════════════════════════════════════════════╗")
        print("║  PHASE 1 — VENDOR DELIVERABLE REVIEW (Primary Workflow)             ║")
        print("║  Agents intake, review, and approve vendor submissions.              ║")
        print("║  Agents do NOT create deliverables in Phase 1.                      ║")
        print("╚══════════════════════════════════════════════════════════════════════╝")
        print()
        print("Phase 1 Commands:")
        print("  intake_deliverable        Intake a vendor-submitted deliverable")
        print("    --file <path>           Path to file (txt, csv, xlsx, etc.)")
        print("    --vendor <name>         Vendor name (e.g. Yubi)")
        print("    --type <type>           Deliverable type (e.g. 'Design Spec')")
        print("  run_deliverable_review    Run all agent reviews on a deliverable")
        print("    --deliverable-id <id>   ID from intake_deliverable")
        print("  deliverable_queue         Show all deliverables and review status")
        print()
        print("Approval & Gate Commands:")
        print("  check_approvals           Show all pending gate approvals")
        print("  resume_workflows          Resume approved workflows")
        print("  process_approval          Record a gate approval response")
        print("    --workflow-id <id>")
        print("    --approver <email>")
        print("    --decision <approved|rejected>")
        print("    --feedback <comment>")
        print()
        print("Budget & Guardrail Commands:")
        print("  budget_status             Show live budget usage report")
        print("  audit_hallucinations      Show recent hallucination flags")
        print()
        print("SME & Learning Commands:")
        print("  add_correction            Record an SME correction for an agent")
        print("    --agent <name>          Agent name (pm_agent, ba_agent, qa_agent, etc.)")
        print("    --hash <output_hash>    Hash of the original output")
        print("    --snippet <text>        Snippet of wrong output")
        print("    --correction <text>     What the correct output should be")
        print("    --category <text>       Correction category")
        print("    --corrector <name>      SME name making the correction")
        print("    --weight <1-5>          Importance weight (default: 1)")
        print("  list_corrections          Show SME corrections for an agent")
        print("    --agent <name>")
        print("  add_frozen_fact           Add a ground-truth fact to guardrails")
        print("    --key <fact_key>")
        print("    --value <fact_value>")
        print("    --added-by <name>")
        print()
        print("Data Intake Commands (BA Agent):")
        print("  run_ba_bdr_intake         Ingest BDR CSV/Excel via BA Agent")
        print("    --file <path>")
        print("  complete_ba_data_analysis Complete analysis after PII clearance")
        print("    --doc-id <id>")
        print("  bdr_queue                 Show BDR document pipeline status")
        print()
        print("Environment & Reporting:")
        print("  env_status                Show environment (lab/production)")
        print("  show_budget_model         Daily budget allocation breakdown")
        print("  show_cost_breakdown       Per-agent token cost breakdown")
        print("  show_optimization_impact  Cost savings analysis")
        print("  project_monthly_cost      Monthly cost scenarios")
        print("  show_executive_summary    Full executive summary")
        print("  token_dashboard           Full token strategy dashboard")
        print("  demo_dashboard            Run dashboard demo")
        sys.exit(1)
        print("\nSME & Guardrail Commands:")
        print("  add_correction            Record an SME correction for an agent")
        print("    --agent <name>          Agent name (pm_agent, ba_agent, etc.)")
        print("    --hash <output_hash>    Hash of the original output")
        print("    --snippet <text>        Snippet of wrong output")
        print("    --correction <text>     What the correct output should be")
        print("    --category <text>       Correction category")
        print("    --corrector <name>      SME name making the correction")
        print("    --weight <1-5>          Importance weight (default: 1)")
        print("  list_corrections          Show SME corrections for an agent")
        print("    --agent <name>")
        print("  add_frozen_fact           Add a new ground-truth fact")
        print("    --key <fact_key>")
        print("    --value <fact_value>")
        print("    --added-by <name>")
        print("\nEnvironment Commands:")
        print("  env_status                Show current environment (lab/production)")
        print("                            Set FG_ENV=lab to switch to lab mode")
        print("\nToken Strategy Commands:")
        print("  show_budget_model         Daily budget allocation")
        print("  show_cost_breakdown       Per-agent cost breakdown (incl. Data Engineer)")
        print("  show_optimization_impact  Technique savings analysis")
        print("  project_monthly_cost      Monthly cost scenarios")
        print("  show_executive_summary    Complete executive summary")
        print("  token_dashboard           Full token strategy dashboard")
        print("\nDashboard & Command Center Commands:")
        print("  demo_dashboard            Run dashboard demo with simulated agents")
        sys.exit(1)

    command = sys.argv[1]

    # ── Dashboard commands (no DB/API needed) ─────────────────────────────────
    if command == "demo_dashboard":
        demo_dashboard(); return

    # ── Token strategy commands (no DB/API needed) ────────────────────────────
    elif command == "show_budget_model":
        show_budget_model(); return
    elif command == "show_cost_breakdown":
        show_cost_breakdown(); return
    elif command == "show_optimization_impact":
        show_optimization_impact(); return
    elif command == "project_monthly_cost":
        show_cost_projections(); return
    elif command == "show_executive_summary":
        show_executive_summary(); return
    elif command == "token_dashboard":
        show_token_dashboard(); return

    # ── Agent commands (require DB + API) ────────────────────────────────────

    agent = AutonomousAgentWithEmailGates()

    # ── PHASE 1: VENDOR DELIVERABLE COMMANDS ─────────────────────────────────

    if command == "intake_deliverable":
        kwargs = {}
        i = 2
        while i < len(sys.argv):
            if sys.argv[i] == "--file"   and i + 1 < len(sys.argv): kwargs["file"]   = sys.argv[i+1]; i += 2
            elif sys.argv[i] == "--vendor" and i + 1 < len(sys.argv): kwargs["vendor"] = sys.argv[i+1]; i += 2
            elif sys.argv[i] == "--type"   and i + 1 < len(sys.argv): kwargs["type"]   = sys.argv[i+1]; i += 2
            else: i += 1
        for req in ("file", "vendor", "type"):
            if req not in kwargs:
                print(f"ERROR: --{req} is required"); sys.exit(1)
        try:
            print(agent.intake_vendor_deliverable(kwargs["file"], kwargs["vendor"], kwargs["type"]))
        except Exception as e:
            logger.error(f"Error: {e}"); print(f"ERROR: {e}")

    elif command == "run_deliverable_review":
        deliverable_id = None
        i = 2
        while i < len(sys.argv):
            if sys.argv[i] == "--deliverable-id" and i + 1 < len(sys.argv):
                deliverable_id = sys.argv[i+1]; i += 2
            else: i += 1
        if not deliverable_id:
            print("ERROR: --deliverable-id is required"); sys.exit(1)
        try:
            print(agent.run_vendor_deliverable_review(deliverable_id))
        except Exception as e:
            logger.error(f"Error: {e}"); print(f"ERROR: {e}")

    elif command == "deliverable_queue":
        try:
            print(agent.get_deliverable_queue())
        except Exception as e:
            logger.error(f"Error: {e}"); print(f"ERROR: {e}")

    elif command == "run_pm_agent":
        logger.info("Running PM Agent with email approval gate...")
        try:
            charter, workflow_state = agent.run_pm_agent_with_gates({
                "client": "Malcolm Goodwin",
                "project": "AURA MVP",
                "timeline_weeks": 12,
                "scope": "Silhouette technology + 3D mesh design + actor model"
            })
            gate = StageGateManager.STAGE_GATES[StageGateName.CHARTER_APPROVAL]
            print("\n" + "=" * 70)
            print(f"WORKFLOW CREATED: {workflow_state.workflow_id}")
            print("=" * 70)
            print(f"Status:      {workflow_state.status.value}")
            print(f"Stage Gate:  {workflow_state.current_stage_gate.value}")
            print(f"Sent To:     {gate.approver_email}")
            print(f"CC:          {', '.join(gate.cc_emails)}")
            print(f"\nGenerated Charter (preview):\n{charter[:300]}...")
            print("\n" + "=" * 70)
            print("Agent paused at stage gate. Awaiting approval email.")
            print("=" * 70)
        except Exception as e:
            logger.error(f"Error: {str(e)}")
            print(f"ERROR: {str(e)}")

    elif command == "check_approvals":
        print(agent.check_pending_approvals())

    elif command == "resume_workflows":
        logger.info("Resuming approved workflows...")
        pending = agent.db.get_approved_workflows_ready_to_resume()
        if not pending:
            print("No approved workflows ready to resume.")
        else:
            for wf in pending:
                result = agent.resume_approved_workflow(wf.workflow_id)
                print(f"\nResumed {wf.workflow_id}\n   Next step: {result}")

    elif command == "process_approval":
        kwargs = {}
        i = 2
        while i < len(sys.argv):
            if sys.argv[i] == "--workflow-id" and i + 1 < len(sys.argv):
                kwargs['workflow_id'] = sys.argv[i + 1]; i += 2
            elif sys.argv[i] == "--approver" and i + 1 < len(sys.argv):
                kwargs['approver_email'] = sys.argv[i + 1]; i += 2
            elif sys.argv[i] == "--decision" and i + 1 < len(sys.argv):
                kwargs['decision'] = sys.argv[i + 1]; i += 2
            elif sys.argv[i] == "--feedback" and i + 1 < len(sys.argv):
                kwargs['feedback'] = sys.argv[i + 1]; i += 2
            else:
                i += 1
        if 'workflow_id' not in kwargs:
            print("ERROR: --workflow-id required")
            sys.exit(1)
        print(agent.process_approval_response(**kwargs))

    elif command == "budget_status":
        print(agent.budget_enforcer.get_status_report())

    elif command == "audit_hallucinations":
        agent.db.cursor.execute(
            "SELECT * FROM hallucination_flags ORDER BY timestamp DESC LIMIT 10"
        )
        flags = agent.db.cursor.fetchall()
        print("\n" + "=" * 70)
        print("HALLUCINATION FLAGS (Last 10)")
        print("=" * 70)
        if not flags:
            print("No hallucination flags detected")
        else:
            for flag_id, agent_name, timestamp, reason, snippet in flags:
                print(f"\n[{flag_id}] {agent_name} @ {timestamp}")
                print(f"  Reason:  {reason}")
                if snippet:
                    print(f"  Snippet: {snippet[:80]}...")

    elif command == "add_correction":
        kwargs = {}
        i = 2
        while i < len(sys.argv):
            if sys.argv[i] == "--agent"      and i + 1 < len(sys.argv): kwargs['agent_name']         = sys.argv[i+1]; i += 2
            elif sys.argv[i] == "--hash"     and i + 1 < len(sys.argv): kwargs['original_output_hash']= sys.argv[i+1]; i += 2
            elif sys.argv[i] == "--snippet"  and i + 1 < len(sys.argv): kwargs['original_snippet']    = sys.argv[i+1]; i += 2
            elif sys.argv[i] == "--correction"and i+1 < len(sys.argv):  kwargs['corrected_content']   = sys.argv[i+1]; i += 2
            elif sys.argv[i] == "--category" and i + 1 < len(sys.argv): kwargs['correction_category'] = sys.argv[i+1]; i += 2
            elif sys.argv[i] == "--corrector"and i + 1 < len(sys.argv): kwargs['corrector_name']      = sys.argv[i+1]; i += 2
            elif sys.argv[i] == "--weight"   and i + 1 < len(sys.argv): kwargs['weight']              = int(sys.argv[i+1]); i += 2
            else: i += 1
        required = ['agent_name', 'original_output_hash', 'original_snippet',
                    'corrected_content', 'correction_category', 'corrector_name']
        missing = [r for r in required if r not in kwargs]
        if missing:
            print(f"ERROR: Missing required args: {missing}")
            sys.exit(1)
        print(agent.submit_sme_correction(**kwargs))

    elif command == "list_corrections":
        kwargs = {}
        i = 2
        while i < len(sys.argv):
            if sys.argv[i] == "--agent" and i + 1 < len(sys.argv):
                kwargs['agent_name'] = sys.argv[i + 1]; i += 2
            else:
                i += 1
        if 'agent_name' not in kwargs:
            print("ERROR: --agent <name> required")
            sys.exit(1)
        corrections = agent.knowledge_library.get_sme_corrections(kwargs['agent_name'], limit=20)
        print(f"\n{'='*70}\nSME CORRECTIONS for {kwargs['agent_name']} ({len(corrections)})\n{'='*70}")
        if not corrections:
            print("No corrections recorded yet.")
        for c in corrections:
            print(f"\n[{c['correction_id'][:8]}] {c['category']} — weight={c['weight']} "
                  f"used={c['usage_count']}x by {c['corrector_name']}")
            print(f"  Wrong:   {c['original_snippet'][:80]}")
            print(f"  Correct: {c['corrected_content'][:80]}")

    elif command == "add_frozen_fact":
        kwargs = {}
        i = 2
        while i < len(sys.argv):
            if sys.argv[i] == "--key"      and i + 1 < len(sys.argv): kwargs['key']      = sys.argv[i+1]; i += 2
            elif sys.argv[i] == "--value"  and i + 1 < len(sys.argv): kwargs['value']    = sys.argv[i+1]; i += 2
            elif sys.argv[i] == "--added-by"and i+1 < len(sys.argv):  kwargs['added_by'] = sys.argv[i+1]; i += 2
            else: i += 1
        if not all(k in kwargs for k in ('key', 'value', 'added_by')):
            print("ERROR: --key, --value, --added-by all required")
            sys.exit(1)
        agent.hallucination_guard.add_frozen_fact(**kwargs)
        print(f"\nFrozen fact added: {kwargs['key']} = {kwargs['value']} (by {kwargs['added_by']})")
        print(f"Active facts: {len(agent.hallucination_guard.FROZEN_FACTS)}")

    elif command == "run_ba_requirements":
        logger.info("Running BA Agent requirements extraction...")
        try:
            result = agent.run_ba_agent_requirements()
            print(result)
        except Exception as e:
            logger.error(f"Error: {str(e)}")
            print(f"ERROR: {str(e)}")

    elif command == "run_ba_bdr_intake":
        filepath = None
        i = 2
        while i < len(sys.argv):
            if sys.argv[i] == "--file" and i + 1 < len(sys.argv):
                filepath = sys.argv[i + 1]; i += 2
            else:
                i += 1
        if not filepath:
            print("ERROR: --file <path> is required")
            sys.exit(1)
        logger.info(f"Running BA BDR intake on: {filepath}")
        try:
            result = agent.run_ba_agent_bdr_intake(filepath)
            print(result)
        except Exception as e:
            logger.error(f"Error: {str(e)}")
            print(f"ERROR: {str(e)}")

    elif command == "complete_ba_data_analysis":
        doc_id = None
        i = 2
        while i < len(sys.argv):
            if sys.argv[i] == "--doc-id" and i + 1 < len(sys.argv):
                doc_id = sys.argv[i + 1]; i += 2
            else:
                i += 1
        if not doc_id:
            print("ERROR: --doc-id <id> is required")
            sys.exit(1)
        logger.info(f"Completing BA data analysis for doc: {doc_id}")
        try:
            result = agent.complete_ba_data_analysis(doc_id)
            print(result)
        except Exception as e:
            logger.error(f"Error: {str(e)}")
            print(f"ERROR: {str(e)}")

    elif command == "bdr_queue":
        try:
            print(agent.get_bdr_queue_status())
        except Exception as e:
            logger.error(f"Error: {str(e)}")
            print(f"ERROR: {str(e)}")

    elif command == "env_status":
        print(f"\n{'='*70}")
        print(f"ENVIRONMENT STATUS")
        print(f"{'='*70}")
        print(f"  FG_ENV:          {FG_ENV}")
        print(f"  IS_LAB_MODE:     {IS_LAB_MODE}")
        print(f"  ACTIVE_MODEL:    {ACTIVE_MODEL}")
        print(f"  Budget Multiplier: {LAB_BUDGET_MULTIPLIER}x")
        if IS_LAB_MODE:
            print(f"\n  [LAB] Emails suppressed, haiku model, relaxed budgets")
            print(f"  Set FG_ENV=production before go-live")
        else:
            print(f"\n  [PRODUCTION] Full guardrails active")
        print(f"{'='*70}\n")

    else:
        print(f"Unknown command: {command}")
        sys.exit(1)

if __name__ == "__main__":
    main()
